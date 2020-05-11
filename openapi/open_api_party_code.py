#-*- coding:utf-8 -*-
import requests
import xmltodict
import json
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.styles import Alignment
import os
"""
open_api 로 실행한 결과 (선거 코드 결과 정보)를 가지고 
해당 선거때 출마한 정당들의 정보를 반환받는다 

미리 저장된 open_api의 결과 선거코드/선거코드.xlsx에서
******선거 ID******   를 가지고 요청함
"""

"""
선거 코드 (선거 날짜 ) 로 정당을 검색함 
결과 값 순서대로
num == 필요없음
sgId == 선거id
jdName == 정당명
pOder == 순서  (기호 번호를 뜻하는듯 ? )
"""
result_file=Workbook()
#도움말 시트 만들기
write_result=result_file.active
write_result.title='도움말'
write_result.merge_cells('A1:H5')

cell = write_result.cell(row=1, column=1)
cell.alignment=Alignment(wrap_text=True)
write_result.cell(row=1,column=1).value='open_api.py로 실행한 결과(선거 코드 결과 정보)를 가지고 \n선거에 참여한 정당을 검색함 (선거코드/선거코드.xlsx)\nsgId = 선거 ID \njdName = 정당명 \npOder = 순서 (기호번호를 뜻하는거같음)'

#엑셀파일 불러오기
load_code=load_workbook('선거코드/선거코드.xlsx',data_only=True)
load_code=load_code['Sheet']
start=0
for i,v in enumerate(load_code.rows):
    if (start==1):
        url='http://apis.data.go.kr/9760000/CommonCodeService/getCommonPartyCodeList'
        data = {
                    'ServiceKey': 'NXjYP6Ks1pRP2JTAbmAel7rQANhWPadkC8BvdLFpmDqVFsMI72sW3ukkWMm3WsXlPmucYg/fxi/WfVJQSLTzkA==',
                    'numOfRows': '1000', 'sgId': ''+str(v[0].value)}
        resp = requests.get(url, params=data)
        # xml형태인 반환값을 json형태로 변환
        request_data = json.dumps(xmltodict.parse(resp.text), indent=4)
        print(request_data)
        request_data = json.loads(request_data)
        request_data['response']['body']['items']['item']=[request_data['response']['body']['items']['item']]
        write_result = result_file.create_sheet(str(v[0].value) + "_" + str(v[1].value) + "_" + str(v[2].value))
        write_result.append(['선거 ID','정당명','순서'])
        for j in request_data['response']['body']['items']['item']:
            write_result.append([j['sgId'],j['jdName'],j['pOrder']])
    elif (v[0].value=='선거 ID'):
        start=1
if not (os.path.isdir('선거참여정당')):
    os.makedirs(os.path.join('선거참여정당'))
result_file.save('선거참여정당/선거참여정당.xlsx')
