import xmltodict
import json
from openpyxl import Workbook
from openpyxl.styles import Alignment
import os
import requests
from urllib.parse import urlencode, quote_plus, unquote
from openpyxl import load_workbook
import pandas as pd

key = 'z20Id2n4sF4V9cqNdIIJNTr13vrmQO9ZZ4R7aLpanlyeoHCuxhiV9eXnYe8g8RwPAHK12gtKzDYYsyObYEwRXw%3D%3D'
url = 'http://apis.data.go.kr/9710000/BillInfoService2/getBillInfoList'
API_Key = unquote(key)

wb = load_workbook(filename='../../국회의원_현황/국회의원.xlsx', data_only=True)
ws = wb.active

kr_name_array = []  # 한글이름 가져오기
hj_name_array = []  # 한자이름 가져오기
bill_no_array = []  # 의안번호 가져오기

for i in range(30, 320):
    kr_name_array.append(ws['C' + str(i)].value)
    hj_name_array.append(ws['E' + str(i)].value)
    bill_no_array.append(ws['B' + str(i)].value)

result_file = Workbook()
write_result = result_file.active
write_result.merge_cells('A3:A17')

write_result.column_dimensions['A'].width = 30

write_result.cell(1, 1, '응답 데이터')

cell = write_result.cell(row=3, column=1)
cell.alignment = Alignment(wrap_text=True)

write_result.cell(row=3, column=1).value = 'billId = 의안ID\n' \
                                           'billNo = 의안번호\n' \
                                           'passGubn = 처리구분\n' \
                                           'billName = 의안명\n' \
                                           'proposalSortation = 발의구분\n' \
                                           'proposalPeople = 공동발의자\n' \
                                           'krName = 한글이름\n' \
                                           'hjName = 한자이름\n' \
                                           'proposerKind = 제안자구분\n' \
                                           'proposeDt = 제안일자\n' \
                                           'procDt = 의결일자\n' \
                                           'generalResult = 의결결과\n' \
                                           'summary = 주요내용\n' \
                                           'procStageCd = 심사진행상태'

num = 24

response_data = ['의안ID',
                 '의안번호',
                 '처리구분',
                 '의안명',
                 '발의구분',
                 '공동발의자',
                 '한글이름',
                 '한자이름',
                 '제안자구분',
                 '제안일자',
                 '의결일자',
                 '의결결과',
                 '주요내용',
                 '심사진행상태']

# B~Z 열의 한 폭길이를 10으로 조정
for i in range(66, 91):
    write_result.column_dimensions['' + (chr(i))].width = 10

for i, str in enumerate(response_data):
    write_result.cell(row=num - 1, column=int(i + 1)).value = str

for i, kr_name in enumerate(kr_name_array):

    # 대표발의 데이터 가져오기
    print('대표발의 : ', len(kr_name_array), ' 번째 중 ', int(i + 1), ' : ' + kr_name + ' 시작')

    # 대표발의
    queryParams = '?' + urlencode(
        {
            quote_plus('pageNo'): '1',
            quote_plus('numOfRows'): 1000,
            quote_plus('mem_name_check'): 'G01',
            quote_plus('mem_name'): kr_name,
            quote_plus('hj_nm'): hj_name_array[i],
            quote_plus('ord'): 'A01',
            quote_plus('start_ord'): 1,
            quote_plus('end_ord'): 999,
            quote_plus('gbn'): 'dae_num_name',
            quote_plus('ServiceKey'): API_Key,
        }
    )

    request = requests.get(url + queryParams)

    # xml형태인 반환값을 json형태로 변환
    request_data = json.dumps(xmltodict.parse(request.text), indent=4)
    request_data = json.loads(request_data)

    for j, v in enumerate(request_data['response']['body']['items']['item']):
        try:
            write_result.cell(row=num + (14 * i) + j, column=1).value = v['billId']
        except:
            pass
        try:
            write_result.cell(row=num + (14 * i) + j, column=2).value = v['billNo']
        except:
            pass
        try:
            write_result.cell(row=num + (14 * i) + j, column=3).value = v['passGubn']
        except:
            pass
        try:
            write_result.cell(row=num + (14 * i) + j, column=4).value = v['billName']
        except:
            pass
        try:
            write_result.cell(row=num + (14 * i) + j, column=5).value = '대표발의'
        except:
            pass
        try:
            write_result.cell(row=num + (14 * i) + j, column=6).value = ''
        except:
            pass
        try:
            write_result.cell(row=num + (14 * i) + j, column=7).value = kr_name
        except:
            pass
        try:
            write_result.cell(row=num + (14 * i) + j, column=8).value = hj_name_array[i]
        except:
            pass
        try:
            write_result.cell(row=num + (14 * i) + j, column=9).value = v['proposerKind']
        except:
            pass
        try:
            write_result.cell(row=num + (14 * i) + j, column=10).value = v['proposeDt']
        except:
            pass
        try:
            write_result.cell(row=num + (14 * i) + j, column=11).value = v['procDt']
        except:
            pass
        try:
            write_result.cell(row=num + (14 * i) + j, column=12).value = v['generalResult']
        except:
            pass
        try:
            write_result.cell(row=num + (14 * i) + j, column=13).value = v['summary']
        except:
            pass
        try:
            write_result.cell(row=num + (14 * i) + j, column=14).value = v['procStageCd']
        except:
            pass
    print('대표발의 : ', len(kr_name_array), ' 번째 중 ', int(i + 1), ' : ' + kr_name + ' 끝\n')

    result_file.save('의안목록검색결과.xlsx')
print("끝")