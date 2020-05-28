# -*- coding: utf-8 -*-

import pymysql as pymysql
import xmltodict
import json
from openpyxl import Workbook
from openpyxl.styles import Alignment
import os
import requests
from urllib.parse import urlencode, quote_plus, unquote
from openpyxl import load_workbook



## 행지우기
# wb = load_workbook(filename='의안목록_대표발의_1인발의_검색결과backup.xlsx', data_only=True)
# ws = wb.active
#
# start_idx = 4895
# end_idx = 8861
#
# index_array = []
#
# for i in range(start_idx, end_idx+1):
#     if (ws['G' + str(i)].value == None ):
#         print('삭제 인덱스 저장 시작 ', i, ' 번째')
#         index_array.append(i)
#
# index_array.reverse()
#
# for idx, remove_index in enumerate(index_array):
#     print('비어있는 행 삭제 시작 ', len(index_array),'중 ', idx, ' 번째')
#     ws.delete_rows(remove_index)
#
# print('끝')
# wb.save('의안목록_대표발의_검색결과.xlsx')

def 디비정보():
    return '52.78.106.225','pmm','tlsehdgnl1!','PoliticsDB','utf8'

host,user,password,db,charset=디비정보()

conn=pymysql.connect(host=host,user=user,password=password,db=db,charset=charset)

curs=conn.cursor()

wb = load_workbook(filename='정당히스토리.xlsx', data_only=True)
ws = wb.active

start_idx = 1
end_idx = 320

index_array = []
당이름 = "민주당@더불어시민당@충청의미래당@복지국가당@가자코리아@남북통일당@민중연합당@무소속@미래한국당@우리당@미래자영업당@기독자유통일당@고용복지연금선진화연대@인권정당@그린불교연합당@우리미래@정의당@대한당@녹색당@사회민주당@직능자영업당@깨어있는시민연대당@개혁국민신당@바른미래당@민주평화당@국민새정당@홍익당@기독당@불교당@친박연대@노동당@한국복지당@통일민주당@가자!평화인권당@늘푸른한국당@한누리평화통일당@자유당@경제애국당@가자환경당@민중민주당@민생당@바른정당@환수복지당@새벽당@자영업당@새누리당@미래통합당@대한민국당@국가혁명배당금당@시대전환@자유한국당@국민행복당@한반도미래연합@공화당@국민희망총연합@한국국민당@국민참여신당@대한애국당@정치개혁연합@여성의당@한나라당@기본소득당@국민대통합당@열린민주당@일제·위안부·인권정당@친반통합@친반통일당@통일한국당@미래당@진리대한당@친박신당@코리아@한국경제당@기독자유당@통합민주당@사이버모바일국민정책당@국민의당@더불어민주당@민중당@친반평화통일당@미래민주당@우리공화당@국제녹색당"

당이름배열 = 당이름.split('@')

추가당배열 = []

for i in range(start_idx, end_idx+1):
    정치인인덱스 = int(ws['A'+str(i)].value)

    정당히스토리정보 = str(ws['C'+str(i)].value)

    정당히스토리정보배열 = 정당히스토리정보.split(',')

    for data in 정당히스토리정보배열:
        시작일 = (data.split('/')[0])
        종료일 = (data.split('/')[1])
        당이름 = (data.split('/')[2])

        추가당배열.append(당이름)

        if 종료일 == '-':
            try:
                당인덱스 = int(당이름배열.index(''+당이름)+1)
                sql = "INSERT INTO PoliticianPartyHistory (politician_idx, party_idx, start_day, end_day, create_at) VALUES (%s, %s, %s,%s,NOW());"
                curs.execute(sql,((정치인인덱스),(당인덱스),(시작일),(None)))
            except:
                당인덱스 = None
                sql = "INSERT INTO PoliticianPartyHistory (politician_idx, party_idx, start_day, end_day, create_at) VALUES (%s, %s, %s,%s,NOW());"
                curs.execute(sql,((정치인인덱스),(당인덱스),(시작일),(None)))
        else:
            try:
                당인덱스 = int(당이름배열.index(''+당이름)+1)
                sql = "INSERT INTO PoliticianPartyHistory (politician_idx, party_idx, start_day, end_day, create_at) VALUES (%s, %s, %s,%s,NOW());"
                curs.execute(sql,((정치인인덱스),(당인덱스),(시작일),(종료일)))
            except:
                당인덱스 = None
                sql = "INSERT INTO PoliticianPartyHistory (politician_idx, party_idx, start_day, end_day, create_at) VALUES (%s, %s, %s,%s,NOW());"
                curs.execute(sql,((정치인인덱스),(당인덱스),(시작일),(종료일)))

    # print(정치인인덱스, 시작일, 종료일, 당이름, 당인덱스)

print(len(추가당배열), list(set(추가당배열)))
print(len(당이름배열), list(set(당이름배열)))

re = []
for i in 추가당배열:
    if i in 당이름배열:
        pass
    else:
        re.append(i)

print(len(re), list(set(re)))

conn.commit()
conn.close()
# 18 32 57  67 84 129 208 276
print('끝')
