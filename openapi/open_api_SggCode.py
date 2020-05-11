import requests
import xmltodict
import json
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.styles import Alignment
import os
"""

선거구 코드 조회

open_api 로 실행한 결과 (선거 코드 결과 정보)를 가지고 
해당 선거때 선거를 실시한 선거구 정보를 반환받는다 

미리 저장된 open_api의 결과 선겈코드/선거코드.xlsx에서
선거ID와 선거종류코드 을 가지고 요청함



반환값은
선거ID 와 선거종류코드를 보고 해당 선거날짜에 선거가 열린 지역을 나타냄

"""
"""
저장되는 값
num == 인덱스 (의미없음)
sgid ==  선거 id
sgTypecode == 선거 종류 
sggName == 선거구명	종로구
sgName == 시도명		서울특별시
wiwName==구시군명 	종로구
sggJungsu == 선출정수 	1
sOrder == 순서  ?
"""

"""
주의사항 : 재·보궐선거의 경우 추가적으로 선거구지역이 추가될 수 있습니다. <<===== 

"""
result_file=Workbook()
#도움말 시트 만들기
write_result=result_file.active
write_result.title='도움말'
write_result.merge_cells('A1:Z10')

cell = write_result.cell(row=1, column=1)
cell.alignment=Alignment(wrap_text=True)
write_result.cell(row=1,column=1).value='open_api.py로 실행한 결과(선거 코드 결과 정보)를 가지고 \n해당선거가 열린 선거구를 검색함 (선거코드/선거코드.xlsx)' \
                                        '\nsgId = 선거 ID \nsgTypecode = 선거종류코드 1.대통형 2.국회의원 3.시도지사 4.구시군장 5.시도의원 6.구시군의회의원 7.국회의원비례대표 8.광역의원비례대표 9.기초의원비례대표 10.교육의원 11.교육감\nsggName = 선거구명\nsgName = 시도명\nwiwName = 구시군명\nsggJungsu = 선출정수\nsOrder = 순서'

#엑셀파일 불러오기
load_code=load_workbook('선거코드/선거코드.xlsx',data_only=True)
load_code=load_code['Sheet']
start=0
for i,v in enumerate(load_code.rows):
    if (start==1 and str(v[2].value) !='0'):
        # print(v[0].value)
        url='http://apis.data.go.kr/9760000/CommonCodeService/getCommonSggCodeList'
        data = {
                    'ServiceKey': 'NXjYP6Ks1pRP2JTAbmAel7rQANhWPadkC8BvdLFpmDqVFsMI72sW3ukkWMm3WsXlPmucYg/fxi/WfVJQSLTzkA==',
                    'numOfRows': '1000', 'sgId': ''+str(v[0].value),'sgTypecode':''+str(v[2].value)}
        resp = requests.get(url, params=data)
        # xml형태인 반환값을 json형태로 변환
        request_data = json.dumps(xmltodict.parse(resp.text), indent=4)
        print(v[0].value,v[2].value)

        request_data = json.loads(request_data)
        write_result = result_file.create_sheet(str(v[0].value) + "_" + str(v[1].value) + "_" + str(v[2].value))
        write_result.append(['선거 ID','선거종류','선거구명','시도명','구시군명','선출정수','순서'])
        print(type(request_data['response']['body']['items']['item']))
        if (str(type(request_data['response']['body']['items']['item']))=="<class 'dict'>"):
            request_data['response']['body']['items']['item']=[request_data['response']['body']['items']['item']]
        # print(request_data)
        for j in request_data['response']['body']['items']['item']:
            write_result.append([str(j['sgId']),str(j['sgTypecode']),str(j['sggName']),str(j['sdName']),str(j['wiwName']),str(j['sggJungsu']),str(j['sOrder'])])
    elif (v[0].value=='선거 ID'):
        start=1
if not (os.path.isdir('선거구코드')):
    os.makedirs(os.path.join('선거구코드'))
result_file.save('선거구코드/선거구코드.xlsx')

