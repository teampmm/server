import xmltodict
import json
from openpyxl import Workbook
from openpyxl.styles import Alignment
import os
import requests
from urllib.parse import urlencode, quote_plus, unquote
from openpyxl import load_workbook

wb = load_workbook(filename='의안목록_대표발의_1인발의_검색결과backup.xlsx', data_only=True)
ws = wb.active

start_idx = 4895
end_idx = 8861

index_array = []

for i in range(start_idx, end_idx+1):
    if (ws['G' + str(i)].value == None ):
        print('삭제 인덱스 저장 시작 ', i, ' 번째')
        index_array.append(i)

index_array.reverse()

for idx, remove_index in enumerate(index_array):
    print('비어있는 행 삭제 시작 ', len(index_array),'중 ', idx, ' 번째')
    ws.delete_rows(remove_index)

print('끝')
wb.save('의안목록_대표발의_검색결과.xlsx')
