#-*- coding:utf-8 -*-
import requests
import xmltodict
import json
from openpyxl import Workbook
from openpyxl.styles import Alignment
import os
"""
선거 코드 조회 api

반환값 결과 (순서대로)
num == 인덱스 (쓸데없음)
sgid == 선거ID
sgName == 선거명
sgTypecode == 선거종류코드 1.대통형 2.국회의원 3.시도지사 4.구시군장 5.시도의원 6.구시군의회의원 7.국회의원비례대표 8.광역의원비례대표 9.기초의원비례대표 10.교육의원 11.교육감
sgVotedate == 선거일자
"""
url = 'http://apis.data.go.kr/9760000/CommonCodeService/getCommonSgCodeList'
data={'ServiceKey':"NXjYP6Ks1pRP2JTAbmAel7rQANhWPadkC8BvdLFpmDqVFsMI72sW3ukkWMm3WsXlPmucYg/fxi/WfVJQSLTzkA==",'numOfRows':'1000'}
resp=requests.get(url,params=data)

#xml형태인 반환값을 json형태로 변환
request_data=json.dumps(xmltodict.parse(resp.text),indent=4)
request_data=json.loads(request_data)

"""
반환값을 엑셀로 저장 하기 위함
"""
result_file=Workbook()
write_result=result_file.active
write_result.merge_cells('A1:H5')

num=13
cell = write_result.cell(row=1, column=1)
cell.alignment=Alignment(wrap_text=True)

write_result.column_dimensions['B'].width=30
write_result.column_dimensions['C'].width=15

write_result.cell(row=1,column=1).value='sgId = 선거 ID \nsgName = 선거명 \nsgTypcode = 선거 종류 코드 (1.대통령 2.국회의원 3.시도지사 4.구시군장 5.시도의원 6.구시군의회의원 7.국회의원비례대표 8.광역의원비례대표 9.기초의원비례대표 10.교육의원 11.교육감)\nsgVotedate = 선거일자'
write_result.cell(row=num-1,column=1).value='선거 ID'
write_result.cell(row=num-1,column=2).value='선거명'
write_result.cell(row=num-1,column=3).value='선거종류코드'
write_result.cell(row=num-1,column=4).value='선거날짜'
for i,v in enumerate(request_data['response']['body']['items']['item']):
    write_result.cell(row=num+i,column=1).value=v['sgId']
    write_result.cell(row=num+i,column=2).value=v['sgName']
    write_result.cell(row=num+i,column=3).value=v['sgTypecode']
    write_result.cell(row=num+i,column=4).value=v['sgVotedate']

if not (os.path.isdir('선거코드')):
    os.makedirs(os.path.join('선거코드'))
result_file.save('선거코드/선거코드.xlsx')
