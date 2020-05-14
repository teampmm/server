import xmltodict
import json
from openpyxl import Workbook
from openpyxl.styles import Alignment
import os
import requests
from urllib.parse import urlencode, quote_plus, unquote
from openpyxl import load_workbook
import pandas as pd

key = 'z20Id2n4sF4V9cqNdIIJNTr13vrmQO9ZZ4R7aLpanlyeoHCuxhiV9eXnYe8g8RwPAHK12gtKzDYYsyObYEwRXw%3D%3D'
url = 'http://apis.data.go.kr/9710000/BillInfoService2/getBillInfoList'
API_Key = unquote(key)

wb = load_workbook(filename='../../국회의원/국회의원.xlsx', data_only=True)
ws = wb.active

kr_name_array = []  # 한글이름 가져오기
hj_name_array = []  # 한자이름 가져오기

no_politician = []

for i in range(30, 320):
    kr_name_array.append(ws['C' + str(i)].value)
    hj_name_array.append(ws['E' + str(i)].value)

# print(kr_name_array.index('황희'))
# print(hj_name_array[kr_name_array.index('황희')])
# exit()

result_file = Workbook()
write_result = result_file.active
write_result.merge_cells('A3:A17')

write_result.column_dimensions['A'].width = 30

write_result.cell(1, 1, '응답 데이터')

cell = write_result.cell(row=3, column=1)
cell.alignment = Alignment(wrap_text=True)

write_result.cell(row=3, column=1).value = 'billId = 의안ID\n' \
                                           'billNo = 의안번호\n' \
                                           'passGubn = 처리구분\n' \
                                           'billName = 의안명\n' \
                                           'proposalSortation = 발의구분\n' \
                                           'proposalPeople = 공동발의자\n' \
                                           'krName = 한글이름\n' \
                                           'hjName = 한자이름\n' \
                                           'proposerKind = 제안자구분\n' \
                                           'proposeDt = 제안일자\n' \
                                           'procDt = 의결일자\n' \
                                           'generalResult = 의결결과\n' \
                                           'summary = 주요내용\n' \
                                           'procStageCd = 심사진행상태'

num = 24
# num = 56024

response_data = ['의안ID',
                 '의안번호',
                 '처리구분',
                 '의안명',
                 '발의구분',
                 '공동발의자',
                 '한글이름',
                 '한자이름',
                 '제안자구분',
                 '제안일자',
                 '의결일자',
                 '의결결과',
                 '주요내용',
                 '심사진행상태']

# B~Z 열의 한 폭길이를 10으로 조정
for i in range(66, 91):
    write_result.column_dimensions['' + (chr(i))].width = 10

for i, st in enumerate(response_data):
    write_result.cell(row=num - 1, column=int(i + 1)).value = st

for count in range(1,58):
    print('57 페이지중 ' , count, ' 페이지 시작')

# quote_plus('pageNo'): count,
# quote_plus('numOfRows'): 1000,
# quote_plus('mem_name_check'): 'G01',
# quote_plus('ord'): 'A01',
# quote_plus('start_ord'): 1,
# quote_plus('end_ord'): 20,
# quote_plus('gbn'): 'dae_num_name',
# quote_plus('ServiceKey'): API_Key

    # 대표발의
    queryParams = '?' + urlencode(
        {
            quote_plus('pageNo'): count,
            quote_plus('numOfRows'): 1000,
            quote_plus('mem_name_check'): 'G01',
            quote_plus('mem_name'): '',
            quote_plus('hj_nm'): '',
            quote_plus('ord'): 'A01',
            quote_plus('start_ord'): 1,
            quote_plus('end_ord'): 20,
            quote_plus('process_num'): '',
            quote_plus('start_process_num'): '',
            quote_plus('end_process_num'): '',
            quote_plus('start_propose_num'): '',
            quote_plus('end_propose_num'): '',
            quote_plus('start_propose_date'): '',
            quote_plus('end_propose_date'): '',
            quote_plus('start_committee_dt'): '',
            quote_plus('end_committee_dt'): '',
            quote_plus('bill_kind_cd'): '',
            quote_plus('curr_committee'): '',
            quote_plus('proposer_kind_cd'): '',
            quote_plus('p_proc_result_cd'): '',
            quote_plus('b_proc_result_cd'): '',
            quote_plus('bill_name'): '',
            quote_plus('gbn'): 'dae_num_name',
            quote_plus('amendmentyn'): '',
            quote_plus('budget'): '',
            quote_plus('ServiceKey'): API_Key
        }
    )

    request = requests.get(url + queryParams)

    # xml형태인 반환값을 json형태로 변환
    request_data = json.dumps(xmltodict.parse(request.text), indent=4)
    request_data = json.loads(request_data)

    # try:
    #     print(kr_name, ' 데이터 : ', len(request_data['response']['body']['items']['item']))
    # except:
    #     pass

    # try:
    #     if (str(type(request_data['response']['body']['items']['item']))=="<class 'dict'>"):
    #         print('데이터가 1개여서 json 형태로 변환')
    #         request_data['response']['body']['items']['item']=[request_data['response']['body']['items']['item']]
    # except:
    #     pass
    print(num , '줄 부터 시작 . 데이터 : ' , len(request_data['response']['body']['items']['item']) , '개')
    for j, v in enumerate(request_data['response']['body']['items']['item']):

        kr_name = str(v['billName'])
        # print('\n',kr_name)
        start_idx = kr_name.rfind('(')
        end_idx = kr_name.rfind('의원')
        kr_name = kr_name[start_idx+1:end_idx]
        # print(kr_name)

        try:
            write_result.cell(row=num + j, column=1).value = v['billId']
        except:
            pass
        try:
            write_result.cell(row=num + j, column=2).value = v['billNo']
        except:
            pass
        try:
            write_result.cell(row=num + j, column=3).value = v['passGubn']
        except:
            pass
        try:
            write_result.cell(row=num + j, column=4).value = v['billName']
        except:
            pass
        try:
            write_result.cell(row=num + j, column=5).value = '대표발의'
        except:
            pass
        try:
            write_result.cell(row=num + j, column=6).value = ''
        except:
            pass
        try:
            write_result.cell(row=num + j, column=7).value = kr_name
        except:
            pass
        try:
            try:
                hj_name = hj_name_array[kr_name_array.index(kr_name)]
                write_result.cell(row=num + j, column=8).value = hj_name
            except:
                no_politician.append(kr_name)
                write_result.cell(row=num + j, column=8).value = ''
        except:
            pass
        try:
            write_result.cell(row=num + j, column=9).value = v['proposerKind']
        except:
            pass
        try:
            write_result.cell(row=num + j, column=10).value = v['proposeDt']
        except:
            pass
        try:
            write_result.cell(row=num + j, column=11).value = v['procDt']
        except:
            pass
        try:
            write_result.cell(row=num + j, column=12).value = v['generalResult']
        except:
            pass
        try:
            write_result.cell(row=num + j, column=13).value = v['summary']
        except:
            pass
        try:
            write_result.cell(row=num + j, column=14).value = v['procStageCd']
        except:
            pass
    num += len(request_data['response']['body']['items']['item'])
    result_file.save('의안목록_대표발의_검색결과.xlsx')

    print('57 페이지중 ' , count, ' 페이지 끝 :  :  끝\n')

print(no_politician)
print("끝")
