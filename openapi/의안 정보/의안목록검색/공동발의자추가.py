import xmltodict
import json
from openpyxl import Workbook
from openpyxl.styles import Alignment
import os
import requests
from urllib.parse import urlencode, quote_plus, unquote
from openpyxl import load_workbook
import pandas as pd
key = 'NXjYP6Ks1pRP2JTAbmAel7rQANhWPadkC8BvdLFpmDqVFsMI72sW3ukkWMm3WsXlPmucYg/fxi/WfVJQSLTzkA=='
url = 'http://apis.data.go.kr/9710000/BillInfoService2/getBillInfoList'
API_Key = unquote(key)

국회의원엑셀 = load_workbook(filename='../../국회의원/국회의원.xlsx', data_only=True)
국회의원 = 국회의원엑셀.active

대표_1인발의엑셀 = load_workbook(filename='의안목록_대표발의_1인발의검색결과.xlsx', data_only=True)
대표_1인발의 = 대표_1인발의엑셀.active

kr_name_array = []  # 한글이름 가져오기
hj_name_array = []  # 한자이름 가져오기
bill_no_array = []  # 의안번호 가져오기

for i in range(30, 320):
    kr_name_array.append(국회의원['AH' + str(i)].value)
    hj_name_array.append(국회의원['AI' + str(i)].value)

for i in range(24, 61373):
    bill_no_array.append(대표_1인발의['B'+ str(i)].value)


print(kr_name_array)
print(hj_name_array)
print(bill_no_array)
no_bill = []

cnt = 0

# for count in range(1,785):
#     print('785 페이지중 ' , count, ' 페이지 시작')
#
#     queryParams = '?' + urlencode(
#         {
#             quote_plus('pageNo'): count,
#             quote_plus('numOfRows'): 1000,
#             quote_plus('mem_name_check'): 'G03',
#             quote_plus('mem_name'): 'G03',
#             quote_plus('hj_nm'): 'G03',
#             quote_plus('ord'): 'A01',
#             quote_plus('start_ord'): 1,
#             quote_plus('end_ord'): 20,
#             quote_plus('gbn'): 'dae_num_name',
#             quote_plus('ServiceKey'): API_Key
#         }
#     )
#     request = requests.get(url + queryParams)
#
#     # xml형태인 반환값을 json형태로 변환
#     request_data = json.dumps(xmltodict.parse(request.text), indent=4)
#     request_data = json.loads(request_data)
#
#
#     try:
#         print('데이터수 : ', len(request_data['response']['body']['items']['item']))
#     except:
#         pass
#
#     for j, v in enumerate(request_data['response']['body']['items']['item']):
#
#         kr_name = str(v['billName'])
#         start_idx = kr_name.rfind('(')
#         end_idx = kr_name.rfind('의원')
#         kr_name = kr_name[start_idx+1:end_idx]
#
#         if(bill_no_array.index(''+v['billNo'])):
#             cnt += 1
#             row = 24 + bill_no_array.index(''+v['billNo'])
#             column = 6
#             공동발의자 = str(대표_1인발의['F' + str(row)].value)
#             if(공동발의자 == None):
#                 공동발의자 = str(kr_name+',')
#             else:
#                 공동발의자 += str(kr_name+',')
#
#             대표_1인발의.cell(row=row, column=column).value = 공동발의자
#         else:
#             pass
#
#     print('공동발의자 : ' , cnt, ' 명 추가')
#     대표_1인발의엑셀.save('공동발의자추가.xlsx')
# print("끝")


for i, kr_name in enumerate(kr_name_array):

    if(i == 5):break

    # 공동발의 데이터 가져오기
    try:
        print('공동발의 : ', len(kr_name_array), ' 번째 중 ', int(i + 1), ' : ' + str(kr_name + ' ') + ' 시작')
    except:
        print('공동발의 : ', len(kr_name_array), ' 번째 중 ', int(i + 1), ' 시작')

    queryParams = '?' + urlencode(
        {
            quote_plus('pageNo'): 1,
            quote_plus('numOfRows'): 5000,
            quote_plus('mem_name_check'): 'G03',
            quote_plus('mem_name'): kr_name_array[i],
            quote_plus('hj_nm'): hj_name_array[i],
            quote_plus('ord'): 'A01',
            quote_plus('start_ord'): 1,
            quote_plus('end_ord'): 20,
            quote_plus('process_num'): '',
            quote_plus('start_process_num'): '',
            quote_plus('end_process_num'): '',
            quote_plus('start_propose_num'): '',
            quote_plus('end_propose_num'): '',
            quote_plus('start_propose_date'): '',
            quote_plus('end_propose_date'): '',
            quote_plus('start_committee_dt'): '',
            quote_plus('end_committee_dt'): '',
            quote_plus('bill_kind_cd'): '',
            quote_plus('curr_committee'): '',
            quote_plus('proposer_kind_cd'): '',
            quote_plus('p_proc_result_cd'): '',
            quote_plus('b_proc_result_cd'): '',
            quote_plus('bill_name'): '',
            quote_plus('gbn'): 'dae_num_name',
            quote_plus('amendmentyn'): '',
            quote_plus('budget'): '',
            quote_plus('ServiceKey'): API_Key
        }
    )
    # 1인발의
    request = requests.get(url + queryParams)

    request_data = json.dumps(xmltodict.parse(request.text), indent=4)
    request_data = json.loads(request_data)

    try:
        print(kr_name, ' 공동발의건수 :  : ', len(request_data['response']['body']['items']['item']))
    except:
        pass

    try:
        if (str(type(request_data['response']['body']['items']['item']))=="<class 'dict'>"):
            print('데이터가 1개여서 json 형태로 변환')
            request_data['response']['body']['items']['item']=[request_data['response']['body']['items']['item']]
    except:
        pass

    cnt = 0
    for j, v in enumerate(request_data['response']['body']['items']['item']):

        if(bill_no_array.index(''+v['billNo'])):
            cnt += 1
            row = 24 + bill_no_array.index(''+v['billNo'])
            column = 6
            공동발의자 = str(대표_1인발의['F' + str(row)].value)
            if(공동발의자 == None):
                공동발의자 = str(kr_name+',')
            else:
                공동발의자 += str(kr_name+',')

            대표_1인발의.cell(row=row, column=column).value = 공동발의자
        else:
            pass
    print('공동발의자 추가 갯수 ', cnt)
    대표_1인발의엑셀.save('공동발의자추가.xlsx')
print("끝")
