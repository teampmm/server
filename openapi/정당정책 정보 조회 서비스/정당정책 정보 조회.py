import xmltodict
import json
from openpyxl import Workbook
from openpyxl.styles import Alignment
import os
import requests
from urllib.parse import urlencode,quote_plus,unquote
from openpyxl import load_workbook

key = 'z20Id2n4sF4V9cqNdIIJNTr13vrmQO9ZZ4R7aLpanlyeoHCuxhiV9eXnYe8g8RwPAHK12gtKzDYYsyObYEwRXw%3D%3D'
url='http://apis.data.go.kr/9760000/PartyPlcInfoInqireService/getPartyPlcInfoInqire'

API_Key=unquote(key)
queryParams='?'+urlencode(
    {
        quote_plus('pageNo'):'1',
        quote_plus('numOfRows'):1000,
        quote_plus('sgId'):20160413,
        quote_plus('ServiceKey'):API_Key,
    }
)

request=requests.get(url+queryParams)

#xml형태인 반환값을 json형태로 변환
request_data=json.dumps(xmltodict.parse(request.text),indent=4)
request_data=json.loads(request_data)

result_file=Workbook()
write_result=result_file.active
write_result.merge_cells('A2:A8')

write_result.column_dimensions['A'].width=30

write_result.cell(1,1,'응답 데이터')

cell = write_result.cell(row=2, column=1)
cell.alignment=Alignment(wrap_text=True)

write_result.cell(row=2,column=1).value='num = 결과순서\n' \
                                        'sgId = 선거ID\n' \
                                        'partyName = 정당명\n' \
                                        'prmsCnt = 공약개수\n' \
                                        'prmsOrd = 공약순번\n' \
                                        'prmsRealmName = 공약분야명\n' \
                                        'prmsTitle = 공약제목명\n' \
                                        'prmmCont = 공약내용'

num = 13

write_result.column_dimensions['B'].width=10
write_result.column_dimensions['C'].width=10
write_result.column_dimensions['D'].width=10
write_result.column_dimensions['E'].width=10
write_result.column_dimensions['F'].width=10
write_result.column_dimensions['G'].width=10
write_result.column_dimensions['H'].width=10


write_result.cell(row=num-1,column=1).value='결과순서'
write_result.cell(row=num-1,column=2).value='선거ID'
write_result.cell(row=num-1,column=3).value='정당명'
write_result.cell(row=num-1,column=4).value='공약개수'
write_result.cell(row=num-1,column=5).value=' '

cnt = 1

for i in range(6,56,5):
    write_result.cell(row=num-1,column=i).value='공약순번'+str(cnt)
    write_result.cell(row=num-1,column=i+1).value='공약분야명'+str(cnt)
    write_result.cell(row=num-1,column=i+2).value='공약제목명'+str(cnt)
    write_result.cell(row=num-1,column=i+3).value='공약내용'+str(cnt)
    write_result.cell(row=num-1,column=i+4).value=' '
    cnt += 1


for i,v in enumerate(request_data['response']['body']['items']['item']):
    write_result.cell(row=num+i,column=1).value=v['num']
    write_result.cell(row=num+i,column=2).value=v['sgId']
    write_result.cell(row=num+i,column=3).value=v['partyName']
    write_result.cell(row=num+i,column=4).value=v['prmsCnt']

for i,v in enumerate(request_data['response']['body']['items']['item']):
    cnt = 1
    for j in range(6, 56, 5):
        write_result.cell(row=num+i,column=j).value=v['prmsOrd'+str(cnt)]
        write_result.cell(row=num+i,column=j+1).value=v['prmsRealmName'+str(cnt)]
        write_result.cell(row=num+i,column=j+2).value=v['prmsTitle'+str(cnt)]
        write_result.cell(row=num+i,column=j+3).value=v['prmmCont'+str(cnt)]
        write_result.cell(row=num+i,column=j+4).value=' '
        cnt += 1

result_file.save('정당정책조회결과.xlsx')
