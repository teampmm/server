#-*- coding:utf-8 -*-
import requests
import xmltodict
import json
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.styles import Alignment
import os

"""
국회의원 현황조회
국회의원 현황 조회 api 를 먼저 실행하고
반환값 deptCd , num 으로 국회의원 상세정보 api 를 실행한다 

그리고 별개로 당선인정보api를 실행해 위에서 실행한 결과값에 이름과 선거구를 비교해서 데이터를 추가함 


"""

def votePolitician(write_result,num,i,k):
    write_result.cell(row=num + i, column=27).value = k['huboid']
    write_result.cell(row=num + i, column=28).value = k['sggName']
    write_result.cell(row=num + i, column=29).value = k['wiwName']
    try:
        write_result.cell(row=num + i, column=30).value = k['wiwName']
    except:
        write_result.cell(row=num + i, column=30).value = ''
    try:
        write_result.cell(row=num + i, column=31).value = k['giho']
    except:
        write_result.cell(row=num + i, column=31).value = ''
    try:
        write_result.cell(row=num + i, column=32).value = k['gihoSangse']
    except:
        write_result.cell(row=num + i, column=32).value = ''
    try:
        write_result.cell(row=num + i, column=33).value = k['jdName']
    except:
        write_result.cell(row=num + i, column=33).value = ''
    try:
        write_result.cell(row=num + i, column=34).value = k['name']
    except:
        write_result.cell(row=num + i, column=34).value = ''
    try:
        write_result.cell(row=num + i, column=35).value = k['hanjaName']
    except:
        write_result.cell(row=num + i, column=35).value = ''
    try:
        write_result.cell(row=num + i, column=36).value = k['gender']
    except:
        write_result.cell(row=num + i, column=36).value = ''
    try:
        write_result.cell(row=num + i, column=37).value = k['birthday']
    except:
        write_result.cell(row=num + i, column=37).value = ''
    try:
        write_result.cell(row=num + i, column=38).value = k['age']
    except:
        write_result.cell(row=num + i, column=38).value = ''
    try:
        write_result.cell(row=num + i, column=39).value = k['addr']
    except:
        write_result.cell(row=num + i, column=39).value = ''
    try:
        write_result.cell(row=num + i, column=40).value = k['jobId']
    except:
        write_result.cell(row=num + i, column=40).value = ''
    try:
        write_result.cell(row=num + i, column=41).value = k['job']
    except:
        write_result.cell(row=num + i, column=41).value = ''
    try:
        write_result.cell(row=num + i, column=42).value = k['eduId']
    except:
        write_result.cell(row=num + i, column=42).value = ''
    try:
        write_result.cell(row=num + i, column=43).value = k['edu']
    except:
        write_result.cell(row=num + i, column=43).value = ''
    try:
        write_result.cell(row=num + i, column=44).value = k['career1']
    except:
        write_result.cell(row=num + i, column=44).value = ''
    try:
        write_result.cell(row=num + i, column=45).value = k['career2']
    except:
        write_result.cell(row=num + i, column=45).value = ''
    try:
        write_result.cell(row=num + i, column=46).value = k['dugsu']
    except:
        write_result.cell(row=num + i, column=46).value = ''
    try:
        write_result.cell(row=num + i, column=47).value = k['dugyul']
    except:
        write_result.cell(row=num + i, column=47).value = ''
    try:
        write_result.cell(row=num + i, column=48).value = k['sgId']
    except:
        write_result.cell(row=num + i, column=48).value = ''
    try:
        write_result.cell(row=num + i, column=49).value = k['sgTypecode']
    except:
        write_result.cell(row=num + i, column=49).value = ''

#당선정보를 가져오기위해 선거코드와 선거종류코드를 가지고 api신청
code_list=[['20160413',4],['20160413',5],['20160413',6],['20160413',7],
           ['20170509',1],['20170509',5],['20170509',6]
           ,['20180613',2],['20180613',3],['20180613',4],['20180613',5],['20180613',6],['20180613',8],['20180613',9],['20180613',10],['20180613',11]
            ,['20190403',2],['20190403',6]]

#당선인 정보조회
url='http://apis.data.go.kr/9760000/WinnerInfoInqireService2/getWinnerInfoInqire'
data = {
            'ServiceKey': 'NXjYP6Ks1pRP2JTAbmAel7rQANhWPadkC8BvdLFpmDqVFsMI72sW3ukkWMm3WsXlPmucYg/fxi/WfVJQSLTzkA==',
            'numOfRows': '1000', 'pageNo':'1','sgId':'20160413','sgTypecode':'2'}
resp = requests.get(url, params=data)
# xml형태인 반환값을 json형태로 변환
vote_result_data = json.dumps(xmltodict.parse(resp.text), indent=4)
vote_result_data = json.loads(vote_result_data)
vote_result_data=vote_result_data['response']['body']['items']['item']
print(vote_result_data)
print(str(len(vote_result_data)))
for i in code_list:
    data = {
        'ServiceKey': 'NXjYP6Ks1pRP2JTAbmAel7rQANhWPadkC8BvdLFpmDqVFsMI72sW3ukkWMm3WsXlPmucYg/fxi/WfVJQSLTzkA==',
        'numOfRows': '1000', 'pageNo': '1', 'sgId': str(i[0]), 'sgTypecode': str(i[1])}
    resp = requests.get(url, params=data)
    # xml형태인 반환값을 json형태로 변환
    vote_data = json.dumps(xmltodict.parse(resp.text), indent=4)
    vote_data = json.loads(vote_data)
    vote_data = vote_data['response']['body']['items']['item']

    for j in vote_data:
        if (j=='num'):
            vote_result_data.append(vote_data)
            break
        else:
            vote_result_data.append(j)
    print(vote_data)
    print(len(vote_result_data))


#국회의원 현황조회
url='http://apis.data.go.kr/9710000/NationalAssemblyInfoService/getMemberCurrStateList'
data = {
            'ServiceKey': 'NXjYP6Ks1pRP2JTAbmAel7rQANhWPadkC8BvdLFpmDqVFsMI72sW3ukkWMm3WsXlPmucYg/fxi/WfVJQSLTzkA==',
            'numOfRows': '1000', 'pageNo':'1'}
resp = requests.get(url, params=data)
# xml형태인 반환값을 json형태로 변환
request_data = json.dumps(xmltodict.parse(resp.text), indent=4)
request_data = json.loads(request_data)

#결과값을 엑셀로 저장하기 위해 엑셀을 실행
result_file=Workbook()
write_result=result_file.active
#셀 병합
write_result.merge_cells('A1:H9')


cell = write_result.cell(row=1, column=1)
cell.alignment=Alignment(wrap_text=True)

num=30

write_result.cell(row=1,column=1).value='국회의원 현황 \ndeptCd = 부서코드 \nnum = 식별코드 \nempNm = 한글이름\nengNm = 영어이름\nhjNm = 한자이름\nreeleGbnNm = 당선횟수\norigNm = 선거구\njpgLink = 의원사진\n' \
                                        '국회의원 상세정보\nempNm = 의원이름\nhjNm = 한자이름\nengNm = 영문이름\nbthDate = 생년월일\npolyNm = 소속정당\norigNm = 선거구\nshrtNm = 소속위원회\nreeleGbnNm = 당선횟수\nelectionNum = 당선대수\n' \
                                        'assemTel = 사무실전화\nassemHomep = 홈페이지\nassemEmail = 이메일\nstaff = 보좌관\nsecretary2 = 비서관2\nsecretary = 비서\nhbbyCd = 취미\nexamCd = 특기\nmemTitle = 약력\n' \
                                        '당선인 정보\nhuboId = 후보자ID\nsggName = 선거구명\nsgName = 시도명\nwiwName = 구시군명\ngiho = 기호\ngihoSangse = 기호상세\njdName = 정당명\nname = 한글성명\nhanjaName = 한자성명\ngender = 성별\nbirthday = 생년월일\nage = 연령\naddr = 주소\njobId = 직업ID\njob = 직업\neduId = 학력ID\nedu = 학력\ncareer1 = 경력1\n' \
                                        'career2 = 경력2\ndugsu = 득표수\ndugyul = 득표율'
#국회의원 현황api
write_result.cell(row=num-1,column=1).value='부서코드'
write_result.cell(row=num-1,column=2).value='식별코드'
write_result.cell(row=num-1,column=3).value='한글이름'
write_result.cell(row=num-1,column=4).value='영어이름'
write_result.cell(row=num-1,column=5).value='한자이름'
write_result.cell(row=num-1,column=6).value='당선횟수'
write_result.cell(row=num-1,column=7).value='선거구'
write_result.cell(row=num-1,column=8).value='의원사진'
#국회의원 상세정보api
write_result.cell(row=num-1,column=9).value='의원이름'
write_result.cell(row=num-1,column=10).value='한자이름'
write_result.cell(row=num-1,column=11).value='영문이름'
write_result.cell(row=num-1,column=12).value='생년월일'
write_result.cell(row=num-1,column=13).value='소속정당'
write_result.cell(row=num-1,column=14).value='선거구'
write_result.cell(row=num-1,column=15).value='소속위원회'
write_result.cell(row=num-1,column=16).value='당선횟수'
write_result.cell(row=num-1,column=17).value='당선대수'
write_result.cell(row=num-1,column=18).value='사무실전화'
write_result.cell(row=num-1,column=19).value='홈페이지'
write_result.cell(row=num-1,column=20).value='이메일'
write_result.cell(row=num-1,column=21).value='보좌관'
write_result.cell(row=num-1,column=22).value='비서관2'
write_result.cell(row=num-1,column=23).value='비서'
write_result.cell(row=num-1,column=24).value='취미'
write_result.cell(row=num-1,column=25).value='특기'
write_result.cell(row=num-1,column=26).value='약력'


#당선일 정보 api
write_result.cell(row=num-1,column=27).value='후보자ID'
write_result.cell(row=num-1,column=28).value='선거구명'
write_result.cell(row=num-1,column=29).value='시도명'
write_result.cell(row=num-1,column=30).value='구시군명'
write_result.cell(row=num-1,column=31).value='기호'
write_result.cell(row=num-1,column=32).value='기호상세'
write_result.cell(row=num-1,column=33).value='정당명'
write_result.cell(row=num-1,column=34).value='한글성명'
write_result.cell(row=num-1,column=35).value='한자성명'
write_result.cell(row=num-1,column=36).value='성별'
write_result.cell(row=num-1,column=37).value='생년월일'
write_result.cell(row=num-1,column=38).value='연령'
write_result.cell(row=num-1,column=39).value='주소'
write_result.cell(row=num-1,column=40).value='직업ID'
write_result.cell(row=num-1,column=41).value='직업'
write_result.cell(row=num-1,column=42).value='학력ID'
write_result.cell(row=num-1,column=43).value='학력'
write_result.cell(row=num-1,column=44).value='경력1'
write_result.cell(row=num-1,column=45).value='경력2'
write_result.cell(row=num-1,column=46).value='득표수'
write_result.cell(row=num-1,column=47).value='득표율'
count = 0
count1 = 0
print(str(len(request_data['response']['body']['items']['item'])))
for i,v in enumerate(request_data['response']['body']['items']['item']):
    engNm=''
    try:
        engNm=v['engNm']
    except:
        pass
    #국회의원 상세조회
    politician_url='http://apis.data.go.kr/9710000/NationalAssemblyInfoService/getMemberDetailInfoList'
    politician_data={
        'ServiceKey':'NXjYP6Ks1pRP2JTAbmAel7rQANhWPadkC8BvdLFpmDqVFsMI72sW3ukkWMm3WsXlPmucYg/fxi/WfVJQSLTzkA==',
        'dept_cd':str(v['deptCd']),
        'num':str(v['num'])

    }
    print(str(v['deptCd']),str(v['num']))
    politician_resp=requests.get(politician_url,params=politician_data)
    politician_request_data = json.dumps(xmltodict.parse(politician_resp.text), indent=4)
    politician_request_data = json.loads(politician_request_data)
    politician_request_data=politician_request_data['response']['body']['item']

    #첫번쨰 요청 == 국회의원 현황
    write_result.cell(row=num+i,column=1).value=v['deptCd']
    write_result.cell(row=num+i,column=2).value=v['num']
    write_result.cell(row=num+i,column=3).value=v['empNm']
    write_result.cell(row=num+i,column=4).value=engNm
    write_result.cell(row=num+i,column=5).value=v['hjNm']
    write_result.cell(row=num+i,column=6).value=v['reeleGbnNm']
    write_result.cell(row=num+i,column=7).value=v['origNm']
    write_result.cell(row=num+i,column=8).value=v['jpgLink']
    #두번쨰요청 == 국회의원 현황의 정보를 가지고 국회의원 상세정보
    write_result.cell(row=num+i,column=9).value=politician_request_data['empNm']
    write_result.cell(row=num+i,column=10).value=politician_request_data['hjNm']
    try:
        write_result.cell(row=num+i,column=11).value=politician_request_data['engNm']
    except:
        write_result.cell(row=num+i,column=11).value=''

    write_result.cell(row=num+i,column=12).value=politician_request_data['bthDate']
    write_result.cell(row=num+i,column=13).value=politician_request_data['polyNm']
    write_result.cell(row=num+i,column=14).value=politician_request_data['origNm']
    try:
        write_result.cell(row=num+i,column=15).value=politician_request_data['shrtNm']
    except:
        write_result.cell(row=num+i,column=15).value=''
    try:
        write_result.cell(row=num+i,column=16).value=politician_request_data['reeleGbnNm']
    except:
        write_result.cell(row=num+i, column=16).value =''
    try:
        write_result.cell(row=num+i,column=17).value=politician_request_data['electionNum']
    except:
        write_result.cell(row=num+i, column=17).value =''
    try:
        write_result.cell(row=num+i,column=18).value=politician_request_data['assemTel']
    except:
        write_result.cell(row=num+i,column=18).value=''
    try:
        write_result.cell(row=num+i,column=19).value=politician_request_data['assemHomep']
    except:
        write_result.cell(row=num+i, column=19).value =''
    try:
        write_result.cell(row=num+i,column=20).value=politician_request_data['assemEmail']
    except:
        write_result.cell(row=num+i,column=20).value=''
    try:
        write_result.cell(row=num+i,column=21).value=politician_request_data['staff']
    except:
        write_result.cell(row=num+i,column=21).value=''
    try:
        write_result.cell(row=num+i,column=22).value=politician_request_data['secretary2']
    except:
        write_result.cell(row=num+i,column=22).value=''
    try:
        write_result.cell(row=num+i,column=23).value=politician_request_data['secretary']
    except:
        write_result.cell(row=num+i,column=23).value=''
    try:
        write_result.cell(row=num+i,column=24).value=politician_request_data['hbbyCd']
    except:
        write_result.cell(row=num+i,column=24).value=''
    try:
        write_result.cell(row=num+i,column=25).value=politician_request_data['examCd']
    except:
        write_result.cell(row=num+i,column=25).value=''
    try:
        write_result.cell(row=num+i,column=26).value=politician_request_data['memTitle']
    except:
        write_result.cell(row=num+i,column=26).value=''
    #당선인 정보 api와 국회의원 상세정보 aspi를 비교해서 데이터를 추가함
    #키값이 없기때문에 이름 , 선거구 를 비교해서 맞는것끼리만 매칭
    for k in vote_result_data:
        # print(str(k['sdName']),str(k['sggName']),v['origNm'],str(k['name']),str(v['empNm']))

        #1. 의원의 (한글,한자)이름이 같다 and 비례대표이다
        #2. 의원의 이름이 같다 and 선거구 이름이 같을때 (서울 동작구 or 전남 OO시 or 세종특별자치시)
        if (str(v['origNm'])=='비례대표' and str(k['sggName'])=='비례대표' and str(k['name'])==str(v['empNm'])):
            #데이터를 담당하는 기관이 달라서 한자 이름이 조금씩 다름 이름이 같고 한자가 2글자 이상 같다면 같은사람으로 표시
            name_check=0
            if (str(k['hanjaName'])[0]==str(v['hjNm'])[0]):
                name_check += 1
            if (str(k['hanjaName'])[1]==str(v['hjNm'])[1]):
                name_check += 1
            if (str(k['hanjaName'])[2]==str(v['hjNm'])[2]):
                name_check += 1
            if(name_check>=2):
                count+=1
                votePolitician(write_result,num,i,k)
                break
        elif(str(k['sdName'])!="전국" and(str(k['name'])==str(v['empNm']) and ((str(k['sdName'])[:2]+" "+str(k['sggName'])==v['origNm']) or (str(k['sdName'])[0]+str(k['sdName'])[2]+" "+str(k['sggName'])==v['origNm']) or (str(k['sdName'])==v['origNm'])))):
            count1+=1
            votePolitician(write_result, num, i, k)
            break
        elif(str(k['sdName'])!="전국" and(str(k['name'])==str(v['empNm']) and str(k['sdName'])[:2]+" "+str(k['sggName'])=='인천 남구갑' and str(v['origNm'])=='인천 미추홀구갑')):
            count1+=1
            votePolitician(write_result, num, i, k)
            break
        elif(str(k['sdName'])!="전국" and(str(k['name'])==str(v['empNm']) and str(k['sdName'])[:2]+" "+str(k['sggName'])=='인천 남구을' and str(v['origNm'])=='인천 미추홀구을')):
            count1+=1
            votePolitician(write_result, num, i, k)
            break

if not (os.path.isdir('국회의원_현황')):
    os.makedirs(os.path.join('국회의원_현황'))
result_file.save('국회의원_현황/국회의원.xlsx')
print(str(count),str(count1))