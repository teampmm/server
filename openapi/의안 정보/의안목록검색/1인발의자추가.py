import xmltodict
import json
from openpyxl import Workbook
from openpyxl.styles import Alignment
import os
import requests
from urllib.parse import urlencode, quote_plus, unquote
from openpyxl import load_workbook
import pandas as pd
key = 'NXjYP6Ks1pRP2JTAbmAel7rQANhWPadkC8BvdLFpmDqVFsMI72sW3ukkWMm3WsXlPmucYg/fxi/WfVJQSLTzkA=='
url = 'http://apis.data.go.kr/9710000/BillInfoService2/getBillInfoList'
API_Key = unquote(key)

wb = load_workbook(filename='../../국회의원/국회의원.xlsx', data_only=True)
ws = wb.active

kr_name_array = []  # 한글이름 가져오기
hj_name_array = []  # 한자이름 가져오기
no_politician = []
for i in range(30, 320):
    kr_name_array.append(ws['AH' + str(i)].value)
    hj_name_array.append(ws['AI' + str(i)].value)

print(kr_name_array)
print(hj_name_array)

wb1 = load_workbook(filename='의안목록_대표발의_검색결과.xlsx', data_only=True)
write_result = wb1.active

# 의안목록_대표발의_검색결과 엑셀파일의 마지막행 + 1
# 4893줄부터 데이터를 추가하겠다 라는 의미
num = 56041

for count in range(1,7):
    print('6 페이지중 ' , count, ' 페이지 시작')

    # quote_plus('pageNo'): count,
    # quote_plus('numOfRows'): 1000,
    # quote_plus('mem_name_check'): 'G02',
    # quote_plus('ord'): 'A01',
    # quote_plus('start_ord'): 1,
    # quote_plus('end_ord'): 20,
    # quote_plus('gbn'): 'dae_num_name',
    # quote_plus('ServiceKey'): API_Key

    queryParams = '?' + urlencode(
        {
            quote_plus('pageNo'): count,
            quote_plus('numOfRows'): 1000,
            quote_plus('mem_name_check'): 'G02',
            quote_plus('mem_name'): '',
            quote_plus('hj_nm'): '',
            quote_plus('ord'): 'A01',
            quote_plus('start_ord'): 1,
            quote_plus('end_ord'): 20,
            quote_plus('process_num'): '',
            quote_plus('start_process_num'): '',
            quote_plus('end_process_num'): '',
            quote_plus('start_propose_num'): '',
            quote_plus('end_propose_num'): '',
            quote_plus('start_propose_date'): '',
            quote_plus('end_propose_date'): '',
            quote_plus('start_committee_dt'): '',
            quote_plus('end_committee_dt'): '',
            quote_plus('bill_kind_cd'): '',
            quote_plus('curr_committee'): '',
            quote_plus('proposer_kind_cd'): '',
            quote_plus('p_proc_result_cd'): '',
            quote_plus('b_proc_result_cd'): '',
            quote_plus('bill_name'): '',
            quote_plus('gbn'): 'dae_num_name',
            quote_plus('amendmentyn'): '',
            quote_plus('budget'): '',
            quote_plus('ServiceKey'): API_Key

        }
    )

    request = requests.get(url + queryParams)

    # xml형태인 반환값을 json형태로 변환
    request_data = json.dumps(xmltodict.parse(request.text), indent=4)
    request_data = json.loads(request_data)

    print(num , '줄 부터 시작 . 데이터 : ' , len(request_data['response']['body']['items']['item']) , '개')
    for j, v in enumerate(request_data['response']['body']['items']['item']):
        kr_name = str(v['billName'])
        start_idx = kr_name.rfind('(')
        end_idx = kr_name.rfind('의원')
        kr_name = kr_name[start_idx+1:end_idx]
        try:
            write_result.cell(row=num + j, column=1).value = v['billId']
        except:
            pass
        try:
            write_result.cell(row=num + j, column=2).value = v['billNo']
        except:
            pass
        try:
            write_result.cell(row=num + j, column=3).value = v['passGubn']
        except:
            pass
        try:
            write_result.cell(row=num + j, column=4).value = v['billName']
        except:
            pass
        try:
            write_result.cell(row=num + j, column=5).value = '1인발의'
        except:
            pass
        try:
            write_result.cell(row=num + j, column=6).value = ''
        except:
            pass
        try:
            write_result.cell(row=num + j, column=7).value = kr_name
        except:
            pass
        try:
            try:
                hj_name = hj_name_array[kr_name_array.index(kr_name)]
                write_result.cell(row=num + j, column=8).value = hj_name
            except:
                no_politician.append(kr_name)
                write_result.cell(row=num + j, column=8).value = ''
        except:
            pass
        try:
            write_result.cell(row=num + j, column=9).value = v['proposerKind']
        except:
            pass
        try:
            write_result.cell(row=num + j, column=10).value = v['proposeDt']
        except:
            pass
        try:
            write_result.cell(row=num + j, column=11).value = v['procDt']
        except:
            pass
        try:
            write_result.cell(row=num + j, column=12).value = v['generalResult']
        except:
            pass
        try:
            write_result.cell(row=num + j, column=13).value = v['summary']
        except:
            pass
        try:
            write_result.cell(row=num + j, column=14).value = v['procStageCd']
        except:
            pass
    num += len(request_data['response']['body']['items']['item'])
    wb1.save('의안목록_대표발의_1인발의검색결과.xlsx')

print("끝")
