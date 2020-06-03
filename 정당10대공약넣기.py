# -*- coding: utf-8 -*-

import pymysql as pymysql
from openpyxl import load_workbook

def 디비정보():
    return '52.78.106.225','pmm','tlsehdgnl1!','PoliticsDB','utf8'

host,user,password,db,charset=디비정보()

conn=pymysql.connect(host=host,user=user,password=password,db=db,charset=charset)

curs=conn.cursor()

wb = load_workbook(filename='party_pledge.xlsx', data_only=True)
ws = wb.active

start_idx = 2
end_idx = 24

curddd = ['AA','AB','AC','AD','AE','AF','AG','AH']

for i in range(start_idx, end_idx+1):

    공약내용 = []
    공약제목 = []

    정당명 = ws['C'+str(i)].value

    공약개수 = (int)(ws['D'+str(i)].value)

    공약최대컬럼 = (68 + 3*공약개수)

    for j in range(69, 공약최대컬럼+1):
        if j > 90:
            컬럼 = curddd[j - 91]
        else:
            컬럼 = chr(j)

        if j % 3 == 2:
            공약내용.append(ws[컬럼+str(i)].value)

        if j % 3 == 1:
            공약제목.append(ws[컬럼+str(i)].value)

    sql = "select idx from Party where name = %s"
    curs.execute(sql,(정당명))
    result = curs.fetchone()
    for idx, data in enumerate(공약내용):
        sql = "INSERT INTO PartyPledge (party_idx, generation, title, content, pledge_implement_status, create_at) VALUES (%s, %s, %s,%s,%s, NOW());"
        curs.execute(sql,((result[0]),(20),(공약제목[idx]),(공약내용[idx]),None))

    print(len(공약제목), 공약제목)
    print(len(공약내용), 공약내용)
    print('\n')

conn.commit()
conn.close()
print('끝')