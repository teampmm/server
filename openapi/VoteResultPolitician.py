import requests
import json
import xmltodict
from openpyxl import Workbook
from openpyxl.styles import Alignment
"""
20160413 에 당선된 지역구 국회의원 + 비례대표
"""
num=20
result_file=Workbook()
write_result=result_file.active
write_result.merge_cells('A1:H9')
cell = write_result.cell(row=1, column=1)
cell.alignment=Alignment(wrap_text=True)
write_result.cell(row=1,column=1).value='num = 인덱스\nsgId = 선거ID\nsgTypecode = 선거종류코드\nhuboId = 후보자ID\nsggName = 선거구명\nsdName = 시도명\nwiwName = 구시군명\ngiho = 기호\ngihoSangse = 기호상세\njdName = 정당명\n' \
                                        'name = 한글성명\nhanjaNmae = 한자성명\ngender = 성별\nbirthdat = 생년월일\nage = 연령\naddr = 주소\njobId = 직업ID\njob = 직업\neduId = 학력ID\nedu = 학력\ncareer1 = 경력1\n' \
                                        'career2 = 경력2\ndugsu = 득표수\ndugyul = 득표율\n'
write_result.cell(row=num,column=1).value='인덱스'
write_result.cell(row=num,column=2).value='선거ID'
write_result.cell(row=num,column=3).value='선거종류코드'
write_result.cell(row=num,column=4).value='후보자ID'
write_result.cell(row=num,column=5).value='선거구명'
write_result.cell(row=num,column=6).value='시도명'
write_result.cell(row=num,column=7).value='구시군명'
write_result.cell(row=num,column=8).value='기호'
write_result.cell(row=num,column=9).value='기호상세'
write_result.cell(row=num,column=10).value='정당명'
write_result.cell(row=num,column=11).value='한글성명'
write_result.cell(row=num,column=12).value='한자성명'
write_result.cell(row=num,column=13).value='성별'
write_result.cell(row=num,column=14).value='생년월일'
write_result.cell(row=num,column=15).value='연령'
write_result.cell(row=num,column=16).value='주소'
write_result.cell(row=num,column=17).value='직업ID'
write_result.cell(row=num,column=18).value='직업'
write_result.cell(row=num,column=19).value='학력ID'
write_result.cell(row=num,column=20).value='학력'
write_result.cell(row=num,column=21).value='경력1'
write_result.cell(row=num,column=22).value='경력2'
write_result.cell(row=num,column=23).value='득표수'
write_result.cell(row=num,column=24).value='득표율'

#당선인 정보조회
url='http://apis.data.go.kr/9760000/WinnerInfoInqireService2/getWinnerInfoInqire'
data = {
            'ServiceKey': '/fxi/==',
            'numOfRows': '1000', 'pageNo':'1','sgId':'20160413','sgTypecode':'2'}
resp = requests.get(url, params=data)
# xml형태인 반환값을 json형태로 변환
vote_result_data = json.dumps(xmltodict.parse(resp.text), indent=4)
vote_result_data = json.loads(vote_result_data)
vote_result_data=vote_result_data['response']['body']['items']['item']

data = {
            'ServiceKey': '/fxi/==',
            'numOfRows': '1000', 'pageNo':'1','sgId':'20160413','sgTypecode':'7'}
resp = requests.get(url, params=data)
vote_result_data_B = json.dumps(xmltodict.parse(resp.text), indent=4)
vote_result_data_B = json.loads(vote_result_data_B)
vote_result_data_B=vote_result_data_B['response']['body']['items']['item']

for i in vote_result_data_B:
    vote_result_data.append(i)
# exit()

for i,v in enumerate(vote_result_data):
    write_result.cell(row=num+i+1, column=1).value = v['num']
    write_result.cell(row=num+i+1, column=2).value = v['sgId']
    write_result.cell(row=num+i+1, column=3).value = v['sgTypecode']
    write_result.cell(row=num+i+1, column=4).value = v['huboid']
    write_result.cell(row=num+i+1, column=5).value = v['sggName']
    write_result.cell(row=num+i+1, column=6).value = v['sdName']
    write_result.cell(row=num+i+1, column=7).value = v['wiwName']
    write_result.cell(row=num+i+1, column=8).value = v['giho']
    write_result.cell(row=num+i+1, column=9).value = v['gihoSangse']
    write_result.cell(row=num+i+1, column=10).value = v['jdName']
    write_result.cell(row=num+i+1, column=11).value = v['name']
    write_result.cell(row=num+i+1, column=12).value = v['hanjaName']
    write_result.cell(row=num+i+1, column=13).value = v['gender']
    write_result.cell(row=num+i+1, column=14).value = v['birthday']
    write_result.cell(row=num+i+1, column=15).value = v['age']
    write_result.cell(row=num+i+1, column=16).value = v['addr']
    write_result.cell(row=num+i+1, column=17).value = v['jobId']
    write_result.cell(row=num+i+1, column=18).value = v['job']
    write_result.cell(row=num+i+1, column=19).value = v['eduId']
    write_result.cell(row=num+i+1, column=20).value = v['edu']
    write_result.cell(row=num+i+1, column=21).value = v['career1']
    write_result.cell(row=num+i+1, column=22).value = v['career2']
    write_result.cell(row=num+i+1, column=23).value = v['dugsu']
    write_result.cell(row=num+i+1, column=24).value = v['dugyul']
print(vote_result_data)
print(str(len(vote_result_data)))
result_file.save('aa.xlsx')
