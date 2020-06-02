""" 대표발의 """
# import xmltodict
# import json
# from openpyxl import Workbook
# from openpyxl.styles import Alignment
# import os
# import requests
# from urllib.parse import urlencode, quote_plus, unquote
# from openpyxl import load_workbook
# import pandas as pd
# import math
#
# # 동휘키
# key = 'z20Id2n4sF4V9cqNdIIJNTr13vrmQO9ZZ4R7aLpanlyeoHCuxhiV9eXnYe8g8RwPAHK12gtKzDYYsyObYEwRXw%3D%3D'
#
# # 종영키
# # key = 'NXjYP6Ks1pRP2JTAbmAel7rQANhWPadkC8BvdLFpmDqVFsMI72sW3ukkWMm3WsXlPmucYg/fxi/WfVJQSLTzkA=='
#
# url = 'http://apis.data.go.kr/9710000/BillInfoService2/getBillInfoList'
# API_Key = unquote(key)
#
# wb = load_workbook(filename='../국회의원이름.xlsx', data_only=True)
# ws = wb.active
#
# kr_name_array = []  # 한글이름 가져오기
# hj_name_array = []  # 한자이름 가져오기
#
# for i in range(1, 321):
#     kr_name_array.append(ws['A' + str(i)].value)
#     hj_name_array.append(ws['B' + str(i)].value)
#
# print(kr_name_array)
# print(hj_name_array)
#
# # 처음할때
# result_file = Workbook()
# #불러올떄
# # result_file = load_workbook(filename='대표1인공동완료.xlsx', data_only=True)
#
# write_result = result_file.active
# write_result.merge_cells('A3:A17')
#
# write_result.column_dimensions['A'].width = 30
#
# write_result.cell(1, 1, '응답 데이터')
#
# cell = write_result.cell(row=3, column=1)
# cell.alignment = Alignment(wrap_text=True)
#
# write_result.cell(row=3, column=1).value = 'billId = 의안ID\n' \
#                                            'billNo = 의안번호\n' \
#                                            'passGubn = 처리구분\n' \
#                                            'billName = 의안명\n' \
#                                            'proposalSortation = 발의구분\n' \
#                                            'proposalPeople = 공동발의자\n' \
#                                            'krName = 한글이름\n' \
#                                            'hjName = 한자이름\n' \
#                                            'proposerKind = 제안자구분\n' \
#                                            'proposeDt = 제안일자\n' \
#                                            'procDt = 의결일자\n' \
#                                            'generalResult = 의결결과\n' \
#                                            'summary = 주요내용\n' \
#                                            'procStageCd = 심사진행상태'
#
# num = 24
# total_count = 0
# response_data = ['의안ID',
#                  '의안번호',
#                  '처리구분',
#                  '의안명',
#                  '발의구분',
#                  '공동발의자',
#                  '한글이름',
#                  '한자이름',
#                  '제안자구분',
#                  '제안일자',
#                  '의결일자',
#                  '의결결과',
#                  '주요내용',
#                  '심사진행상태']
#
# # B~Z 열의 한 폭길이를 10으로 조정
# for i in range(66, 91):
#     write_result.column_dimensions['' + (chr(i))].width = 10
#
# # response data 넣기
# for i, str in enumerate(response_data):
#     write_result.cell(row=num - 1, column=int(i + 1)).value = str
#
#
# # G01 - 대표발의
# # G02 - 1인발의
# # G03 - 공동발의
#
# for idx,K_NAME in enumerate(kr_name_array):
#
#     # if K_NAME != "심재철":
#     #     continue
#
#     print(idx, '번쨰 : ',K_NAME,' 시작', num,'행 부터 시작')
#
#     if K_NAME == "김성태" or K_NAME == "최경환":
#         # 대표발의
#         queryParams = '?' + urlencode(
#             {
#                 quote_plus('pageNo'): 1,
#                 quote_plus('numOfRows'): 1000,
#                 quote_plus('mem_name_check'): 'G01',
#                 quote_plus('mem_name'): '',
#                 quote_plus('hj_nm'): hj_name_array[idx],
#                 quote_plus('ord'): 'A01',
#                 quote_plus('start_ord'): 20,
#                 quote_plus('end_ord'): 20,
#                 quote_plus('process_num'): '',
#                 quote_plus('start_process_num'): '',
#                 quote_plus('end_process_num'): '',
#                 quote_plus('propose_num'): '',
#                 quote_plus('start_propose_num'): '',
#                 quote_plus('end_propose_num'): '',
#                 quote_plus('start_propose_date'): '',
#                 quote_plus('end_propose_date'): '',
#                 quote_plus('start_committee_dt'): '',
#                 quote_plus('end_committee_dt'): '',
#                 quote_plus('bill_kind_cd'): '',
#                 quote_plus('curr_committee'): '',
#                 quote_plus('proposer_kind_cd'): '',
#                 quote_plus('p_proc_result_cd'): '',
#                 quote_plus('b_proc_result_cd'): '',
#                 quote_plus('bill_name'): '',
#                 quote_plus('amendmentyn'): '',
#                 quote_plus('budget'): '',
#                 quote_plus('gbn'): 'dae_num_name',
#                 quote_plus('ServiceKey'): API_Key,
#             }
#         )
#     else:
#         # 대표발의
#         queryParams = '?' + urlencode(
#             {
#                 quote_plus('pageNo'): 1,
#                 quote_plus('numOfRows'): 1000,
#                 quote_plus('mem_name_check'): 'G01',
#                 quote_plus('mem_name'): K_NAME,
#                 quote_plus('hj_nm'): '',
#                 quote_plus('ord'): 'A01',
#                 quote_plus('start_ord'): 20,
#                 quote_plus('end_ord'): 20,
#                 quote_plus('process_num'): '',
#                 quote_plus('start_process_num'): '',
#                 quote_plus('end_process_num'): '',
#                 quote_plus('propose_num'): '',
#                 quote_plus('start_propose_num'): '',
#                 quote_plus('end_propose_num'): '',
#                 quote_plus('start_propose_date'): '',
#                 quote_plus('end_propose_date'): '',
#                 quote_plus('start_committee_dt'): '',
#                 quote_plus('end_committee_dt'): '',
#                 quote_plus('bill_kind_cd'): '',
#                 quote_plus('curr_committee'): '',
#                 quote_plus('proposer_kind_cd'): '',
#                 quote_plus('p_proc_result_cd'): '',
#                 quote_plus('b_proc_result_cd'): '',
#                 quote_plus('bill_name'): '',
#                 quote_plus('amendmentyn'): '',
#                 quote_plus('budget'): '',
#                 quote_plus('gbn'): 'dae_num_name',
#                 quote_plus('ServiceKey'): API_Key,
#             }
#         )
#
#     request = requests.get(url + queryParams)
#     # xml형태인 반환값을 json형태로 변환
#     request_data = json.dumps(xmltodict.parse(request.text), indent=4)
#     request_data = json.loads(request_data)
#
#     # print(request_data)
#
#     print('대표 발의 갯수 : ', request_data['response']['body']['totalCount'])
#     page = int(math.ceil(int(request_data['response']['body']['totalCount'])/1000))
#
#       try:
#           if (str(type(request_data['response']['body']['items']['item']))=="<class 'dict'>"):
#               print('데이터가 1개여서 json 형태로 변환')
#               request_data['response']['body']['items']['item']=[request_data['response']['body']['items']['item']]
#       except:
#           pass
#
#     if page == 1:
#         for j, v in enumerate(request_data['response']['body']['items']['item']):
#             try:
#                 write_result.cell(row=num + j, column=1).value = v['billId']
#             except:
#                 pass
#             try:
#                 write_result.cell(row=num + j, column=2).value = v['billNo']
#             except:
#                 pass
#             try:
#                 write_result.cell(row=num + j, column=3).value = v['passGubn']
#             except:
#                 pass
#             try:
#                 write_result.cell(row=num + j, column=4).value = v['billName']
#             except:
#                 pass
#             try:
#                 write_result.cell(row=num + j, column=5).value = '대표발의'
#             except:
#                 pass
#             try:
#                 write_result.cell(row=num + j, column=6).value = ''
#             except:
#                 pass
#             try:
#                 write_result.cell(row=num + j, column=7).value = K_NAME
#             except:
#                 pass
#             try:
#                 write_result.cell(row=num + j, column=8).value = hj_name_array[idx]
#             except:
#                 pass
#             try:
#                 write_result.cell(row=num + j, column=9).value = v['proposerKind']
#             except:
#                 pass
#             try:
#                 write_result.cell(row=num + j, column=10).value = v['proposeDt']
#             except:
#                 pass
#             try:
#                 write_result.cell(row=num + j, column=11).value = v['procDt']
#             except:
#                 pass
#             try:
#                 write_result.cell(row=num + j, column=12).value = v['generalResult']
#             except:
#                 pass
#             try:
#                 write_result.cell(row=num + j, column=13).value = v['summary']
#             except:
#                 pass
#             try:
#                 write_result.cell(row=num + j, column=14).value = v['procStageCd']
#             except:
#                 pass
#
#     else:
#         for p in range(1, page+1):
#
#             if K_NAME == "김성태" or K_NAME == "최경환":
#                     # 대표발의
#                 queryParams = '?' + urlencode(
#                     {
#                         quote_plus('pageNo'): p,
#                         quote_plus('numOfRows'): 1000,
#                         quote_plus('mem_name_check'): 'G01',
#                         quote_plus('mem_name'): '',
#                         quote_plus('hj_nm'): hj_name_array[idx],
#                         quote_plus('ord'): 'A01',
#                         quote_plus('start_ord'): 20,
#                         quote_plus('end_ord'): 20,
#                         quote_plus('process_num'): '',
#                         quote_plus('start_process_num'): '',
#                         quote_plus('end_process_num'): '',
#                         quote_plus('propose_num'): '',
#                         quote_plus('start_propose_num'): '',
#                         quote_plus('end_propose_num'): '',
#                         quote_plus('start_propose_date'): '',
#                         quote_plus('end_propose_date'): '',
#                         quote_plus('start_committee_dt'): '',
#                         quote_plus('end_committee_dt'): '',
#                         quote_plus('bill_kind_cd'): '',
#                         quote_plus('curr_committee'): '',
#                         quote_plus('proposer_kind_cd'): '',
#                         quote_plus('p_proc_result_cd'): '',
#                         quote_plus('b_proc_result_cd'): '',
#                         quote_plus('bill_name'): '',
#                         quote_plus('amendmentyn'): '',
#                         quote_plus('budget'): '',
#                         quote_plus('gbn'): 'dae_num_name',
#                         quote_plus('ServiceKey'): API_Key,
#                     }
#                 )
#             else:
#                 # 대표발의
#                 queryParams = '?' + urlencode(
#                     {
#                         quote_plus('pageNo'): p,
#                         quote_plus('numOfRows'): 1000,
#                         quote_plus('mem_name_check'): 'G01',
#                         quote_plus('mem_name'): K_NAME,
#                         quote_plus('hj_nm'): '',
#                         quote_plus('ord'): 'A01',
#                         quote_plus('start_ord'): 20,
#                         quote_plus('end_ord'): 20,
#                         quote_plus('process_num'): '',
#                         quote_plus('start_process_num'): '',
#                         quote_plus('end_process_num'): '',
#                         quote_plus('propose_num'): '',
#                         quote_plus('start_propose_num'): '',
#                         quote_plus('end_propose_num'): '',
#                         quote_plus('start_propose_date'): '',
#                         quote_plus('end_propose_date'): '',
#                         quote_plus('start_committee_dt'): '',
#                         quote_plus('end_committee_dt'): '',
#                         quote_plus('bill_kind_cd'): '',
#                         quote_plus('curr_committee'): '',
#                         quote_plus('proposer_kind_cd'): '',
#                         quote_plus('p_proc_result_cd'): '',
#                         quote_plus('b_proc_result_cd'): '',
#                         quote_plus('bill_name'): '',
#                         quote_plus('amendmentyn'): '',
#                         quote_plus('budget'): '',
#                         quote_plus('gbn'): 'dae_num_name',
#                         quote_plus('ServiceKey'): API_Key,
#                     }
#                 )
#
#             request = requests.get(url + queryParams)
#             # xml형태인 반환값을 json형태로 변환
#             request_data = json.dumps(xmltodict.parse(request.text), indent=4)
#             request_data = json.loads(request_data)
#
#             for j, v in enumerate(request_data['response']['body']['items']['item']):
#                 try:
#                     write_result.cell(row=num + j, column=1).value = v['billId']
#                 except:
#                     pass
#                 try:
#                     write_result.cell(row=num + j, column=2).value = v['billNo']
#                 except:
#                     pass
#                 try:
#                     write_result.cell(row=num + j, column=3).value = v['passGubn']
#                 except:
#                     pass
#                 try:
#                     write_result.cell(row=num + j, column=4).value = v['billName']
#                 except:
#                     pass
#                 try:
#                     write_result.cell(row=num + j, column=5).value = '대표발의'
#                 except:
#                     pass
#                 try:
#                     write_result.cell(row=num + j, column=6).value = ''
#                 except:
#                     pass
#                 try:
#                     write_result.cell(row=num + j, column=7).value = K_NAME
#                 except:
#                     pass
#                 try:
#                     write_result.cell(row=num + j, column=8).value = hj_name_array[idx]
#                 except:
#                     pass
#                 try:
#                     write_result.cell(row=num + j, column=9).value = v['proposerKind']
#                 except:
#                     pass
#                 try:
#                     write_result.cell(row=num + j, column=10).value = v['proposeDt']
#                 except:
#                     pass
#                 try:
#                     write_result.cell(row=num + j, column=11).value = v['procDt']
#                 except:
#                     pass
#                 try:
#                     write_result.cell(row=num + j, column=12).value = v['generalResult']
#                 except:
#                     pass
#                 try:
#                     write_result.cell(row=num + j, column=13).value = v['summary']
#                 except:
#                     pass
#                 try:
#                     write_result.cell(row=num + j, column=14).value = v['procStageCd']
#                 except:
#                     pass
#
#     num += int(request_data['response']['body']['totalCount'])
#     total_count += int(request_data['response']['body']['totalCount'])
#     result_file.save('대표1인공동완료.xlsx')
#     print(idx, '번쨰 저장완료 : ',K_NAME,' 끝\n')
#
# print(total_count, "개 끝")

""" 1인발의 """
# import xmltodict
# import json
# from openpyxl import Workbook
# from openpyxl.styles import Alignment
# import os
# import requests
# from urllib.parse import urlencode, quote_plus, unquote
# from openpyxl import load_workbook
# import pandas as pd
# import math
#
# # 동휘키
# key = 'z20Id2n4sF4V9cqNdIIJNTr13vrmQO9ZZ4R7aLpanlyeoHCuxhiV9eXnYe8g8RwPAHK12gtKzDYYsyObYEwRXw%3D%3D'
#
# # 종영키
# # key = 'NXjYP6Ks1pRP2JTAbmAel7rQANhWPadkC8BvdLFpmDqVFsMI72sW3ukkWMm3WsXlPmucYg/fxi/WfVJQSLTzkA=='
#
# url = 'http://apis.data.go.kr/9710000/BillInfoService2/getBillInfoList'
# API_Key = unquote(key)
#
# wb = load_workbook(filename='../국회의원이름.xlsx', data_only=True)
# ws = wb.active
#
# kr_name_array = []  # 한글이름 가져오기
# hj_name_array = []  # 한자이름 가져오기
#
# for i in range(1, 321):
#     kr_name_array.append(ws['A' + str(i)].value)
#     hj_name_array.append(ws['B' + str(i)].value)
#
# print(kr_name_array)
# print(hj_name_array)
#
# # 처음할때
# # result_file = Workbook()
# #불러올떄
# result_file = load_workbook(filename='대표1인공동완료.xlsx', data_only=True)
#
# write_result = result_file.active
#
# num = 21962
# total_count = 0
# # G01 - 대표발의
# # G02 - 1인발의
# # G03 - 공동발의
#
# for idx,K_NAME in enumerate(kr_name_array):
#
#     # if K_NAME != "정세균":
#     #     continue
#
#     print(idx, '번쨰 : ',K_NAME,' 시작', num,'행 부터 시작')
#
#     if K_NAME == "김성태" or K_NAME == "최경환":
#         # 1인발의
#         queryParams = '?' + urlencode(
#             {
#                 quote_plus('pageNo'): 1,
#                 quote_plus('numOfRows'): 1000,
#                 quote_plus('mem_name_check'): 'G02',
#                 quote_plus('mem_name'): '',
#                 quote_plus('hj_nm'): hj_name_array[idx],
#                 quote_plus('ord'): 'A01',
#                 quote_plus('start_ord'): 20,
#                 quote_plus('end_ord'): 20,
#                 quote_plus('process_num'): '',
#                 quote_plus('start_process_num'): '',
#                 quote_plus('end_process_num'): '',
#                 quote_plus('propose_num'): '',
#                 quote_plus('start_propose_num'): '',
#                 quote_plus('end_propose_num'): '',
#                 quote_plus('start_propose_date'): '',
#                 quote_plus('end_propose_date'): '',
#                 quote_plus('start_committee_dt'): '',
#                 quote_plus('end_committee_dt'): '',
#                 quote_plus('bill_kind_cd'): '',
#                 quote_plus('curr_committee'): '',
#                 quote_plus('proposer_kind_cd'): '',
#                 quote_plus('p_proc_result_cd'): '',
#                 quote_plus('b_proc_result_cd'): '',
#                 quote_plus('bill_name'): '',
#                 quote_plus('amendmentyn'): '',
#                 quote_plus('budget'): '',
#                 quote_plus('gbn'): 'dae_num_name',
#                 quote_plus('ServiceKey'): API_Key,
#             }
#         )
#     else:
#         # 1인발의
#         queryParams = '?' + urlencode(
#             {
#                 quote_plus('pageNo'): 1,
#                 quote_plus('numOfRows'): 1000,
#                 quote_plus('mem_name_check'): 'G02',
#                 quote_plus('mem_name'): K_NAME,
#                 quote_plus('hj_nm'): '',
#                 quote_plus('ord'): 'A01',
#                 quote_plus('start_ord'): 20,
#                 quote_plus('end_ord'): 20,
#                 quote_plus('process_num'): '',
#                 quote_plus('start_process_num'): '',
#                 quote_plus('end_process_num'): '',
#                 quote_plus('propose_num'): '',
#                 quote_plus('start_propose_num'): '',
#                 quote_plus('end_propose_num'): '',
#                 quote_plus('start_propose_date'): '',
#                 quote_plus('end_propose_date'): '',
#                 quote_plus('start_committee_dt'): '',
#                 quote_plus('end_committee_dt'): '',
#                 quote_plus('bill_kind_cd'): '',
#                 quote_plus('curr_committee'): '',
#                 quote_plus('proposer_kind_cd'): '',
#                 quote_plus('p_proc_result_cd'): '',
#                 quote_plus('b_proc_result_cd'): '',
#                 quote_plus('bill_name'): '',
#                 quote_plus('amendmentyn'): '',
#                 quote_plus('budget'): '',
#                 quote_plus('gbn'): 'dae_num_name',
#                 quote_plus('ServiceKey'): API_Key,
#             }
#         )
#
#     request = requests.get(url + queryParams)
#     # xml형태인 반환값을 json형태로 변환
#     request_data = json.dumps(xmltodict.parse(request.text), indent=4)
#     request_data = json.loads(request_data)
#
#     # print(request_data)
#
#     print('1인 발의 갯수 : ', request_data['response']['body']['totalCount'])
#     if int(request_data['response']['body']['totalCount']) == 0:
#         print('발의개수 0개 패스')
#         continue
#
#     page = int(math.ceil(int(request_data['response']['body']['totalCount'])/1000))
#
#     try:
#         if (str(type(request_data['response']['body']['items']['item']))=="<class 'dict'>"):
#             print('데이터가 1개여서 json 형태로 변환')
#             request_data['response']['body']['items']['item']=[request_data['response']['body']['items']['item']]
#     except:
#         pass
#
#     if page == 1:
#         for j, v in enumerate(request_data['response']['body']['items']['item']):
#             try:
#                 write_result.cell(row=num + j, column=1).value = v['billId']
#             except:
#                 pass
#             try:
#                 write_result.cell(row=num + j, column=2).value = v['billNo']
#             except:
#                 pass
#             try:
#                 write_result.cell(row=num + j, column=3).value = v['passGubn']
#             except:
#                 pass
#             try:
#                 write_result.cell(row=num + j, column=4).value = v['billName']
#             except:
#                 pass
#             try:
#                 write_result.cell(row=num + j, column=5).value = '1인발의'
#             except:
#                 pass
#             try:
#                 write_result.cell(row=num + j, column=6).value = ''
#             except:
#                 pass
#             try:
#                 write_result.cell(row=num + j, column=7).value = K_NAME
#             except:
#                 pass
#             try:
#                 write_result.cell(row=num + j, column=8).value = hj_name_array[idx]
#             except:
#                 pass
#             try:
#                 write_result.cell(row=num + j, column=9).value = v['proposerKind']
#             except:
#                 pass
#             try:
#                 write_result.cell(row=num + j, column=10).value = v['proposeDt']
#             except:
#                 pass
#             try:
#                 write_result.cell(row=num + j, column=11).value = v['procDt']
#             except:
#                 pass
#             try:
#                 write_result.cell(row=num + j, column=12).value = v['generalResult']
#             except:
#                 pass
#             try:
#                 write_result.cell(row=num + j, column=13).value = v['summary']
#             except:
#                 pass
#             try:
#                 write_result.cell(row=num + j, column=14).value = v['procStageCd']
#             except:
#                 pass
#
#     else:
#         for p in range(1, page+1):
#
#             if K_NAME == "김성태" or K_NAME == "최경환":
#                 # 대표발의
#                 queryParams = '?' + urlencode(
#                     {
#                         quote_plus('pageNo'): p,
#                         quote_plus('numOfRows'): 1000,
#                         quote_plus('mem_name'): '',
#                         quote_plus('mem_name_check'): 'G02',
#                         quote_plus('hj_nm'): hj_name_array[idx],
#                         quote_plus('ord'): 'A01',
#                         quote_plus('start_ord'): 20,
#                         quote_plus('end_ord'): 20,
#                         quote_plus('process_num'): '',
#                         quote_plus('start_process_num'): '',
#                         quote_plus('end_process_num'): '',
#                         quote_plus('propose_num'): '',
#                         quote_plus('start_propose_num'): '',
#                         quote_plus('end_propose_num'): '',
#                         quote_plus('start_propose_date'): '',
#                         quote_plus('end_propose_date'): '',
#                         quote_plus('start_committee_dt'): '',
#                         quote_plus('end_committee_dt'): '',
#                         quote_plus('bill_kind_cd'): '',
#                         quote_plus('curr_committee'): '',
#                         quote_plus('proposer_kind_cd'): '',
#                         quote_plus('p_proc_result_cd'): '',
#                         quote_plus('b_proc_result_cd'): '',
#                         quote_plus('bill_name'): '',
#                         quote_plus('amendmentyn'): '',
#                         quote_plus('budget'): '',
#                         quote_plus('gbn'): 'dae_num_name',
#                         quote_plus('ServiceKey'): API_Key,
#                     }
#                 )
#             else:
#                 # 대표발의
#                 queryParams = '?' + urlencode(
#                     {
#                         quote_plus('pageNo'): p,
#                         quote_plus('numOfRows'): 1000,
#                         quote_plus('mem_name_check'): 'G02',
#                         quote_plus('mem_name'): K_NAME,
#                         quote_plus('hj_nm'): '',
#                         quote_plus('ord'): 'A01',
#                         quote_plus('start_ord'): 20,
#                         quote_plus('end_ord'): 20,
#                         quote_plus('process_num'): '',
#                         quote_plus('start_process_num'): '',
#                         quote_plus('end_process_num'): '',
#                         quote_plus('propose_num'): '',
#                         quote_plus('start_propose_num'): '',
#                         quote_plus('end_propose_num'): '',
#                         quote_plus('start_propose_date'): '',
#                         quote_plus('end_propose_date'): '',
#                         quote_plus('start_committee_dt'): '',
#                         quote_plus('end_committee_dt'): '',
#                         quote_plus('bill_kind_cd'): '',
#                         quote_plus('curr_committee'): '',
#                         quote_plus('proposer_kind_cd'): '',
#                         quote_plus('p_proc_result_cd'): '',
#                         quote_plus('b_proc_result_cd'): '',
#                         quote_plus('bill_name'): '',
#                         quote_plus('amendmentyn'): '',
#                         quote_plus('budget'): '',
#                         quote_plus('gbn'): 'dae_num_name',
#                         quote_plus('ServiceKey'): API_Key,
#                     }
#                 )
#
#             request = requests.get(url + queryParams)
#             # xml형태인 반환값을 json형태로 변환
#             request_data = json.dumps(xmltodict.parse(request.text), indent=4)
#             request_data = json.loads(request_data)
#
#             for j, v in enumerate(request_data['response']['body']['items']['item']):
#                 try:
#                     write_result.cell(row=num + j, column=1).value = v['billId']
#                 except:
#                     pass
#                 try:
#                     write_result.cell(row=num + j, column=2).value = v['billNo']
#                 except:
#                     pass
#                 try:
#                     write_result.cell(row=num + j, column=3).value = v['passGubn']
#                 except:
#                     pass
#                 try:
#                     write_result.cell(row=num + j, column=4).value = v['billName']
#                 except:
#                     pass
#                 try:
#                     write_result.cell(row=num + j, column=5).value = '1인발의'
#                 except:
#                     pass
#                 try:
#                     write_result.cell(row=num + j, column=6).value = ''
#                 except:
#                     pass
#                 try:
#                     write_result.cell(row=num + j, column=7).value = K_NAME
#                 except:
#                     pass
#                 try:
#                     write_result.cell(row=num + j, column=8).value = hj_name_array[idx]
#                 except:
#                     pass
#                 try:
#                     write_result.cell(row=num + j, column=9).value = v['proposerKind']
#                 except:
#                     pass
#                 try:
#                     write_result.cell(row=num + j, column=10).value = v['proposeDt']
#                 except:
#                     pass
#                 try:
#                     write_result.cell(row=num + j, column=11).value = v['procDt']
#                 except:
#                     pass
#                 try:
#                     write_result.cell(row=num + j, column=12).value = v['generalResult']
#                 except:
#                     pass
#                 try:
#                     write_result.cell(row=num + j, column=13).value = v['summary']
#                 except:
#                     pass
#                 try:
#                     write_result.cell(row=num + j, column=14).value = v['procStageCd']
#                 except:
#                     pass
#
#     num += int(request_data['response']['body']['totalCount'])
#     total_count += int(request_data['response']['body']['totalCount'])
#     result_file.save('대표1인공동완료.xlsx')
#     print(idx, '번째 저장완료 : ',K_NAME,' 끝\n\n')
#
# print(total_count, "개 끝")

# """ 공동발의 추가 """
# import xmltodict
# import json
# from openpyxl import Workbook
# from openpyxl.styles import Alignment
# import os
# import requests
# from urllib.parse import urlencode, quote_plus, unquote
# from openpyxl import load_workbook
# import pandas as pd
# import math
#
# # 동휘키
# key = 'z20Id2n4sF4V9cqNdIIJNTr13vrmQO9ZZ4R7aLpanlyeoHCuxhiV9eXnYe8g8RwPAHK12gtKzDYYsyObYEwRXw%3D%3D'
#
# # 종영키
# # key = 'NXjYP6Ks1pRP2JTAbmAel7rQANhWPadkC8BvdLFpmDqVFsMI72sW3ukkWMm3WsXlPmucYg/fxi/WfVJQSLTzkA=='
#
# url = 'http://apis.data.go.kr/9710000/BillInfoService2/getBillInfoList'
# API_Key = unquote(key)
#
# wb = load_workbook(filename='../국회의원이름.xlsx', data_only=True)
# ws = wb.active
#
# kr_name_array = []  # 한글이름 가져오기
# hj_name_array = []  # 한자이름 가져오기
#
# for i in range(1, 321):
#     kr_name_array.append(ws['A' + str(i)].value)
#     hj_name_array.append(ws['B' + str(i)].value)
#
# print(kr_name_array)
# print(hj_name_array)
#
# # 처음할때
# # result_file = Workbook()
# #불러올떄
# result_file = load_workbook(filename='대표1인공동완료.xlsx', data_only=True)
#
# write_result = result_file.active
# write_result.merge_cells('A3:A17')
#
# write_result.column_dimensions['A'].width = 30
#
# write_result.cell(1, 1, '응답 데이터')
#
# cell = write_result.cell(row=3, column=1)
# cell.alignment = Alignment(wrap_text=True)
#
# write_result.cell(row=3, column=1).value = 'billId = 의안ID\n' \
#                                            'billNo = 의안번호\n' \
#                                            'passGubn = 처리구분\n' \
#                                            'billName = 의안명\n' \
#                                            'proposalSortation = 발의구분\n' \
#                                            'proposalPeople = 공동발의자\n' \
#                                            'krName = 한글이름\n' \
#                                            'hjName = 한자이름\n' \
#                                            'proposerKind = 제안자구분\n' \
#                                            'proposeDt = 제안일자\n' \
#                                            'procDt = 의결일자\n' \
#                                            'generalResult = 의결결과\n' \
#                                            'summary = 주요내용\n' \
#                                            'procStageCd = 심사진행상태'
#
# num = 21905
#
# response_data = ['의안ID',
#                  '의안번호',
#                  '처리구분',
#                  '의안명',
#                  '발의구분',
#                  '공동발의자',
#                  '한글이름',
#                  '한자이름',
#                  '제안자구분',
#                  '제안일자',
#                  '의결일자',
#                  '의결결과',
#                  '주요내용',
#                  '심사진행상태']
#
# # B~Z 열의 한 폭길이를 10으로 조정
# for i in range(66, 91):
#     write_result.column_dimensions['' + (chr(i))].width = 10
#
# # # response data 넣기
# # for i, str in enumerate(response_data):
# #     write_result.cell(row=num - 1, column=int(i + 1)).value = str
#
#
# # G01 - 대표발의
# # G02 - 1인발의
# # G03 - 공동발의
#
# for idx,K_NAME in enumerate(kr_name_array):
#
#     # if K_NAME != "정세균":
#     #     continue
#
#     print(idx, '번쨰 : ',K_NAME,' 시작', num,'행 부터 시작')
#
#     if K_NAME == "김성태" or K_NAME == "최경환":
#         # 1인발의
#         queryParams = '?' + urlencode(
#             {
#                 quote_plus('pageNo'): 1,
#                 quote_plus('numOfRows'): 1000,
#                 quote_plus('mem_name_check'): 'G02',
#                 quote_plus('mem_name'): '',
#                 quote_plus('hj_nm'): hj_name_array[idx],
#                 quote_plus('ord'): 'A01',
#                 quote_plus('start_ord'): 20,
#                 quote_plus('end_ord'): 20,
#                 quote_plus('process_num'): '',
#                 quote_plus('start_process_num'): '',
#                 quote_plus('end_process_num'): '',
#                 quote_plus('propose_num'): '',
#                 quote_plus('start_propose_num'): '',
#                 quote_plus('end_propose_num'): '',
#                 quote_plus('start_propose_date'): '',
#                 quote_plus('end_propose_date'): '',
#                 quote_plus('start_committee_dt'): '',
#                 quote_plus('end_committee_dt'): '',
#                 quote_plus('bill_kind_cd'): '',
#                 quote_plus('curr_committee'): '',
#                 quote_plus('proposer_kind_cd'): '',
#                 quote_plus('p_proc_result_cd'): '',
#                 quote_plus('b_proc_result_cd'): '',
#                 quote_plus('bill_name'): '',
#                 quote_plus('amendmentyn'): '',
#                 quote_plus('budget'): '',
#                 quote_plus('gbn'): 'dae_num_name',
#                 quote_plus('ServiceKey'): API_Key,
#             }
#         )
#     else:
#         # 1인발의
#         queryParams = '?' + urlencode(
#             {
#                 quote_plus('pageNo'): 1,
#                 quote_plus('numOfRows'): 1000,
#                 quote_plus('mem_name_check'): 'G02',
#                 quote_plus('mem_name'): K_NAME,
#                 quote_plus('hj_nm'): '',
#                 quote_plus('ord'): 'A01',
#                 quote_plus('start_ord'): 20,
#                 quote_plus('end_ord'): 20,
#                 quote_plus('process_num'): '',
#                 quote_plus('start_process_num'): '',
#                 quote_plus('end_process_num'): '',
#                 quote_plus('propose_num'): '',
#                 quote_plus('start_propose_num'): '',
#                 quote_plus('end_propose_num'): '',
#                 quote_plus('start_propose_date'): '',
#                 quote_plus('end_propose_date'): '',
#                 quote_plus('start_committee_dt'): '',
#                 quote_plus('end_committee_dt'): '',
#                 quote_plus('bill_kind_cd'): '',
#                 quote_plus('curr_committee'): '',
#                 quote_plus('proposer_kind_cd'): '',
#                 quote_plus('p_proc_result_cd'): '',
#                 quote_plus('b_proc_result_cd'): '',
#                 quote_plus('bill_name'): '',
#                 quote_plus('amendmentyn'): '',
#                 quote_plus('budget'): '',
#                 quote_plus('gbn'): 'dae_num_name',
#                 quote_plus('ServiceKey'): API_Key,
#             }
#         )
#
#     request = requests.get(url + queryParams)
#     # xml형태인 반환값을 json형태로 변환
#     request_data = json.dumps(xmltodict.parse(request.text), indent=4)
#     request_data = json.loads(request_data)
#
#     # print(request_data)
#
#     print('1인 발의 갯수 : ', request_data['response']['body']['totalCount'])
#     page = int(math.ceil(int(request_data['response']['body']['totalCount'])/1000))
#
#     try:
#         if (str(type(request_data['response']['body']['items']['item']))=="<class 'dict'>"):
#             print('데이터가 1개여서 json 형태로 변환')
#             request_data['response']['body']['items']['item']=[request_data['response']['body']['items']['item']]
#     except:
#         pass
#
#     if page == 1:
#         for j, v in enumerate(request_data['response']['body']['items']['item']):
#             try:
#                 write_result.cell(row=num + j, column=1).value = v['billId']
#             except:
#                 pass
#             try:
#                 write_result.cell(row=num + j, column=2).value = v['billNo']
#             except:
#                 pass
#             try:
#                 write_result.cell(row=num + j, column=3).value = v['passGubn']
#             except:
#                 pass
#             try:
#                 write_result.cell(row=num + j, column=4).value = v['billName']
#             except:
#                 pass
#             try:
#                 write_result.cell(row=num + j, column=5).value = '1인발의'
#             except:
#                 pass
#             try:
#                 write_result.cell(row=num + j, column=6).value = ''
#             except:
#                 pass
#             try:
#                 write_result.cell(row=num + j, column=7).value = K_NAME
#             except:
#                 pass
#             try:
#                 write_result.cell(row=num + j, column=8).value = hj_name_array[idx]
#             except:
#                 pass
#             try:
#                 write_result.cell(row=num + j, column=9).value = v['proposerKind']
#             except:
#                 pass
#             try:
#                 write_result.cell(row=num + j, column=10).value = v['proposeDt']
#             except:
#                 pass
#             try:
#                 write_result.cell(row=num + j, column=11).value = v['procDt']
#             except:
#                 pass
#             try:
#                 write_result.cell(row=num + j, column=12).value = v['generalResult']
#             except:
#                 pass
#             try:
#                 write_result.cell(row=num + j, column=13).value = v['summary']
#             except:
#                 pass
#             try:
#                 write_result.cell(row=num + j, column=14).value = v['procStageCd']
#             except:
#                 pass
#
#     else:
#         for p in range(1, page+1):
#
#             if K_NAME == "김성태" or K_NAME == "최경환":
#                 # 대표발의
#                 queryParams = '?' + urlencode(
#                     {
#                         quote_plus('pageNo'): p,
#                         quote_plus('numOfRows'): 1000,
#                         quote_plus('mem_name'): '',
#                         quote_plus('mem_name_check'): 'G02',
#                         quote_plus('hj_nm'): hj_name_array[idx],
#                         quote_plus('ord'): 'A01',
#                         quote_plus('start_ord'): 20,
#                         quote_plus('end_ord'): 20,
#                         quote_plus('process_num'): '',
#                         quote_plus('start_process_num'): '',
#                         quote_plus('end_process_num'): '',
#                         quote_plus('propose_num'): '',
#                         quote_plus('start_propose_num'): '',
#                         quote_plus('end_propose_num'): '',
#                         quote_plus('start_propose_date'): '',
#                         quote_plus('end_propose_date'): '',
#                         quote_plus('start_committee_dt'): '',
#                         quote_plus('end_committee_dt'): '',
#                         quote_plus('bill_kind_cd'): '',
#                         quote_plus('curr_committee'): '',
#                         quote_plus('proposer_kind_cd'): '',
#                         quote_plus('p_proc_result_cd'): '',
#                         quote_plus('b_proc_result_cd'): '',
#                         quote_plus('bill_name'): '',
#                         quote_plus('amendmentyn'): '',
#                         quote_plus('budget'): '',
#                         quote_plus('gbn'): 'dae_num_name',
#                         quote_plus('ServiceKey'): API_Key,
#                     }
#                 )
#             else:
#                 # 대표발의
#                 queryParams = '?' + urlencode(
#                     {
#                         quote_plus('pageNo'): p,
#                         quote_plus('numOfRows'): 1000,
#                         quote_plus('mem_name_check'): 'G02',
#                         quote_plus('mem_name'): K_NAME,
#                         quote_plus('hj_nm'): '',
#                         quote_plus('ord'): 'A01',
#                         quote_plus('start_ord'): 20,
#                         quote_plus('end_ord'): 20,
#                         quote_plus('process_num'): '',
#                         quote_plus('start_process_num'): '',
#                         quote_plus('end_process_num'): '',
#                         quote_plus('propose_num'): '',
#                         quote_plus('start_propose_num'): '',
#                         quote_plus('end_propose_num'): '',
#                         quote_plus('start_propose_date'): '',
#                         quote_plus('end_propose_date'): '',
#                         quote_plus('start_committee_dt'): '',
#                         quote_plus('end_committee_dt'): '',
#                         quote_plus('bill_kind_cd'): '',
#                         quote_plus('curr_committee'): '',
#                         quote_plus('proposer_kind_cd'): '',
#                         quote_plus('p_proc_result_cd'): '',
#                         quote_plus('b_proc_result_cd'): '',
#                         quote_plus('bill_name'): '',
#                         quote_plus('amendmentyn'): '',
#                         quote_plus('budget'): '',
#                         quote_plus('gbn'): 'dae_num_name',
#                         quote_plus('ServiceKey'): API_Key,
#                     }
#                 )
#
#             request = requests.get(url + queryParams)
#             # xml형태인 반환값을 json형태로 변환
#             request_data = json.dumps(xmltodict.parse(request.text), indent=4)
#             request_data = json.loads(request_data)
#
#             for j, v in enumerate(request_data['response']['body']['items']['item']):
#                 try:
#                     write_result.cell(row=num + j, column=1).value = v['billId']
#                 except:
#                     pass
#                 try:
#                     write_result.cell(row=num + j, column=2).value = v['billNo']
#                 except:
#                     pass
#                 try:
#                     write_result.cell(row=num + j, column=3).value = v['passGubn']
#                 except:
#                     pass
#                 try:
#                     write_result.cell(row=num + j, column=4).value = v['billName']
#                 except:
#                     pass
#                 try:
#                     write_result.cell(row=num + j, column=5).value = '1인발의'
#                 except:
#                     pass
#                 try:
#                     write_result.cell(row=num + j, column=6).value = ''
#                 except:
#                     pass
#                 try:
#                     write_result.cell(row=num + j, column=7).value = K_NAME
#                 except:
#                     pass
#                 try:
#                     write_result.cell(row=num + j, column=8).value = hj_name_array[idx]
#                 except:
#                     pass
#                 try:
#                     write_result.cell(row=num + j, column=9).value = v['proposerKind']
#                 except:
#                     pass
#                 try:
#                     write_result.cell(row=num + j, column=10).value = v['proposeDt']
#                 except:
#                     pass
#                 try:
#                     write_result.cell(row=num + j, column=11).value = v['procDt']
#                 except:
#                     pass
#                 try:
#                     write_result.cell(row=num + j, column=12).value = v['generalResult']
#                 except:
#                     pass
#                 try:
#                     write_result.cell(row=num + j, column=13).value = v['summary']
#                 except:
#                     pass
#                 try:
#                     write_result.cell(row=num + j, column=14).value = v['procStageCd']
#                 except:
#                     pass
#
#     num += int(request_data['response']['body']['totalCount'])
#
#     result_file.save('대표1인공동완료.xlsx')
#     print(idx, '번째 저장완료 : ',K_NAME,' 끝\n')
#
# print("끝")

""" 공동발의 추가 """

import xmltodict
import json
from openpyxl import Workbook
from openpyxl.styles import Alignment
import os
import requests
from urllib.parse import urlencode, quote_plus, unquote
from openpyxl import load_workbook
import pandas as pd
import math

# 동휘키
key = 'z20Id2n4sF4V9cqNdIIJNTr13vrmQO9ZZ4R7aLpanlyeoHCuxhiV9eXnYe8g8RwPAHK12gtKzDYYsyObYEwRXw%3D%3D'

# 종영키
# key = 'NXjYP6Ks1pRP2JTAbmAel7rQANhWPadkC8BvdLFpmDqVFsMI72sW3ukkWMm3WsXlPmucYg/fxi/WfVJQSLTzkA=='

url = 'http://apis.data.go.kr/9710000/BillInfoService2/getBillInfoList'
API_Key = unquote(key)

wb = load_workbook(filename='../국회의원이름.xlsx', data_only=True)
ws = wb.active

tete = load_workbook(filename='대표_1인발의완료.xlsx', data_only=True)
te = tete.active

kr_name_array = []  # 한글이름 가져오기
hj_name_array = []  # 한자이름 가져오기
bill_no_array = []  # 한자이름 가져오기

# 법안배열
bill_content_array = []
bill_content_idx = []

for i in range(1, 321):
    kr_name_array.append(ws['A' + str(i)].value)
    hj_name_array.append(ws['B' + str(i)].value)


for i in range(24, 22585):
    value = str(te['B' + str(i)].value)
    bill_no_array.append(value)

# print(type(bill_no_array[18195]))
# print(type(bill_no_array[18194]))
# exit()
# print(kr_name_array)
# print(hj_name_array)
# print(len(bill_no_array))

#불러올떄
result_file = load_workbook(filename='의안목록검색결과.xlsx', data_only=True)

write_result = result_file.active
total_count = 0
# G01 - 대표발의
# G02 - 1인발의
# G03 - 공동발의

for idx,K_NAME in enumerate(kr_name_array):

    # if idx < 141:
    #     continue

    print(idx, '번쨰 : ',K_NAME,' 시작')

    if K_NAME == "김성태" or K_NAME == "최경환":
        print('동명이인')
        # 1인발의
        queryParams = '?' + urlencode(
            {
                quote_plus('pageNo'): 1,
                quote_plus('numOfRows'): 1000,
                quote_plus('mem_name_check'): 'G03',
                quote_plus('mem_name'): '',
                quote_plus('hj_nm'): hj_name_array[idx],
                quote_plus('ord'): 'A01',
                quote_plus('start_ord'): 20,
                quote_plus('end_ord'): 20,
                quote_plus('process_num'): '',
                quote_plus('start_process_num'): '',
                quote_plus('end_process_num'): '',
                quote_plus('propose_num'): '',
                quote_plus('start_propose_num'): '',
                quote_plus('end_propose_num'): '',
                quote_plus('start_propose_date'): '',
                quote_plus('end_propose_date'): '',
                quote_plus('start_committee_dt'): '',
                quote_plus('end_committee_dt'): '',
                quote_plus('bill_kind_cd'): '',
                quote_plus('curr_committee'): '',
                quote_plus('proposer_kind_cd'): '',
                quote_plus('p_proc_result_cd'): '',
                quote_plus('b_proc_result_cd'): '',
                quote_plus('bill_name'): '',
                quote_plus('amendmentyn'): '',
                quote_plus('budget'): '',
                quote_plus('gbn'): 'dae_num_name',
                quote_plus('ServiceKey'): API_Key,
            }
        )
    else:
        # 1인발의
        queryParams = '?' + urlencode(
            {
                quote_plus('pageNo'): 1,
                quote_plus('numOfRows'): 1000,
                quote_plus('mem_name_check'): 'G03',
                quote_plus('mem_name'): K_NAME,
                quote_plus('hj_nm'): '',
                quote_plus('ord'): 'A01',
                quote_plus('start_ord'): 20,
                quote_plus('end_ord'): 20,
                quote_plus('process_num'): '',
                quote_plus('start_process_num'): '',
                quote_plus('end_process_num'): '',
                quote_plus('propose_num'): '',
                quote_plus('start_propose_num'): '',
                quote_plus('end_propose_num'): '',
                quote_plus('start_propose_date'): '',
                quote_plus('end_propose_date'): '',
                quote_plus('start_committee_dt'): '',
                quote_plus('end_committee_dt'): '',
                quote_plus('bill_kind_cd'): '',
                quote_plus('curr_committee'): '',
                quote_plus('proposer_kind_cd'): '',
                quote_plus('p_proc_result_cd'): '',
                quote_plus('b_proc_result_cd'): '',
                quote_plus('bill_name'): '',
                quote_plus('amendmentyn'): '',
                quote_plus('budget'): '',
                quote_plus('gbn'): 'dae_num_name',
                quote_plus('ServiceKey'): API_Key,
            }
        )

    request = requests.get(url + queryParams)
    # xml형태인 반환값을 json형태로 변환
    request_data = json.dumps(xmltodict.parse(request.text), indent=4)
    request_data = json.loads(request_data)

    print('공동 발의 갯수 : ', request_data['response']['body']['totalCount'])
    if int(request_data['response']['body']['totalCount']) == 0:
        print('발의개수 0개 패스')
        continue

    page = int(math.ceil(int(request_data['response']['body']['totalCount'])/1000))
    print('총 ', page, ' 페이지')
    try:
        if (str(type(request_data['response']['body']['items']['item']))=="<class 'dict'>"):
            print('데이터가 1개여서 json 형태로 변환')
            request_data['response']['body']['items']['item']=[request_data['response']['body']['items']['item']]
    except:
        pass

    if page == 1:
        for j, v in enumerate(request_data['response']['body']['items']['item']):

            bill_no = str((v['billNo']))

            # 법안번호 인덱스 찾기

            bill_idx = bill_no_array.index(bill_no) + 24

            # bill_together = str(te['F' + str(bill_idx)].value)

            # bill_together += K_NAME + ','

            if K_NAME == "김성태" or K_NAME == "최경환":

                name = K_NAME + "(" + hj_name_array[idx] + ")"
                print(name)

                bill_content_array.append([bill_idx,str(name+",")])
                bill_content_idx.append(bill_idx)
            else:
                bill_content_array.append([bill_idx,str(K_NAME+",")])
                bill_content_idx.append(bill_idx)

            # try:
            #     write_result.cell(row=bill_idx, column=6).value = bill_together
            # except:
            #     pass

    else:
        for p in range(1, page+1):
            print(p,'페이지 시작')
            if K_NAME == "김성태" or K_NAME == "최경환":
                # 공동발의
                queryParams = '?' + urlencode(
                    {
                        quote_plus('pageNo'): p,
                        quote_plus('numOfRows'): 1000,
                        quote_plus('mem_name'): '',
                        quote_plus('mem_name_check'): 'G03',
                        quote_plus('hj_nm'): hj_name_array[idx],
                        quote_plus('ord'): 'A01',
                        quote_plus('start_ord'): 20,
                        quote_plus('end_ord'): 20,
                        quote_plus('process_num'): '',
                        quote_plus('start_process_num'): '',
                        quote_plus('end_process_num'): '',
                        quote_plus('propose_num'): '',
                        quote_plus('start_propose_num'): '',
                        quote_plus('end_propose_num'): '',
                        quote_plus('start_propose_date'): '',
                        quote_plus('end_propose_date'): '',
                        quote_plus('start_committee_dt'): '',
                        quote_plus('end_committee_dt'): '',
                        quote_plus('bill_kind_cd'): '',
                        quote_plus('curr_committee'): '',
                        quote_plus('proposer_kind_cd'): '',
                        quote_plus('p_proc_result_cd'): '',
                        quote_plus('b_proc_result_cd'): '',
                        quote_plus('bill_name'): '',
                        quote_plus('amendmentyn'): '',
                        quote_plus('budget'): '',
                        quote_plus('gbn'): 'dae_num_name',
                        quote_plus('ServiceKey'): API_Key,
                    }
                )
            else:
                # 공동발의
                queryParams = '?' + urlencode(
                    {
                        quote_plus('pageNo'): p,
                        quote_plus('numOfRows'): 1000,
                        quote_plus('mem_name_check'): 'G03',
                        quote_plus('mem_name'): K_NAME,
                        quote_plus('hj_nm'): '',
                        quote_plus('ord'): 'A01',
                        quote_plus('start_ord'): 20,
                        quote_plus('end_ord'): 20,
                        quote_plus('process_num'): '',
                        quote_plus('start_process_num'): '',
                        quote_plus('end_process_num'): '',
                        quote_plus('propose_num'): '',
                        quote_plus('start_propose_num'): '',
                        quote_plus('end_propose_num'): '',
                        quote_plus('start_propose_date'): '',
                        quote_plus('end_propose_date'): '',
                        quote_plus('start_committee_dt'): '',
                        quote_plus('end_committee_dt'): '',
                        quote_plus('bill_kind_cd'): '',
                        quote_plus('curr_committee'): '',
                        quote_plus('proposer_kind_cd'): '',
                        quote_plus('p_proc_result_cd'): '',
                        quote_plus('b_proc_result_cd'): '',
                        quote_plus('bill_name'): '',
                        quote_plus('amendmentyn'): '',
                        quote_plus('budget'): '',
                        quote_plus('gbn'): 'dae_num_name',
                        quote_plus('ServiceKey'): API_Key,
                    }
                )

            request = requests.get(url + queryParams)
            # xml형태인 반환값을 json형태로 변환
            request_data = json.dumps(xmltodict.parse(request.text), indent=4)
            request_data = json.loads(request_data)

            for j, v in enumerate(request_data['response']['body']['items']['item']):

                bill_no = str(v['billNo'])

                # 법안번호 인덱스 찾기

                bill_idx = bill_no_array.index(bill_no) + 24

                # bill_together = str(te['F' + str(bill_idx)].value)

                # bill_together += K_NAME + ','

                if K_NAME == "김성태" or K_NAME == "최경환":

                    name = K_NAME + "(" + hj_name_array[idx] + ")"
                    print(name)
                    bill_content_array.append([bill_idx,str(name+",")])
                    bill_content_idx.append(bill_idx)
                else:
                    bill_content_array.append([bill_idx,str(K_NAME+",")])
                    bill_content_idx.append(bill_idx)

                # try:
                #     write_result.cell(row=bill_idx, column=6).value = bill_together
                # except:
                #     pass

    total_count += int(request_data['response']['body']['totalCount'])

    print(idx, '번째 저장완료 : ',K_NAME,' 끝\n\n')


bill_content_idx = list(set(bill_content_idx))

cnt = 0

for i in bill_content_idx:
    strr = ""
    for j in bill_content_array:
        if j[0] == i:
            cnt += 1
            strr += str(j[1])
    write_result.cell(row=i, column=6).value = strr

result_file.save('0601대표1인공동완료.xlsx')

print(total_count, "개 끝 ", cnt)